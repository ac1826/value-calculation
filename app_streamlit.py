# -*- coding: utf-8 -*-
import streamlit as st
import pandas as pd
import numpy as np
import io
import re
import unicodedata as ud
from pathlib import Path
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

# —— 指定物料号统一 ×1.09（数量不变）——
CODES_FACTOR_109 = {
    "ADB0100000048", "ADB0100362848", "ADS0300000048", "ADL0400000048",
    "ADL2000000048", "ADL2018362848", "ADL0343362848", "ADL2054362848",
    "ADL0359362848", "ADT0100362848", "ADD0120000048", "ADO0100000048",
    "ADO0600000048", "ADB0700000048", "ADN0600362848", "ADO0700362848",
    "ADO0600362848", "ADB0600000048", "ADG0800000048", "ADG1100002848",
    "ADL0900362848", "ADL0251000028", "ABL0169332340", "ADB1600362848",
    "ADF0100000048", "ADG1900493648", "ADW0800000448", "ADW0843000048",
    "ADB1600005618", "ADG1300000018", "ADB0200000018", "ADB0100000018",
    "ADB0100366510", "ADS0300000018", "ADL0400000018", "ADL0463364418",
    "ADL2000000018", "ADL0455000418", "ADL2059000418", "ADL0454000418",
    "ADD0120000018", "ADO0100000018", "ADV0100000018", "ADB0700000018",
    "ADN0500000018", "ADB0600000018", "ADG0800000018", "ADN0151000018",
    "ABC0192233300", "ADL0459000048", "ADB0700000058", "ADB0600000058",
}





def tax_factor_for_code(code: str) -> float:
    """命中名单 → 1.09；否则 1.0"""
    return 1.09 if str(code).strip() in CODES_FACTOR_109 else 1.0


def resolve_sel_month(overview):
    """Return the selected month Period('M') robustly using sel/sel2/overview; never raises NameError."""
    import pandas as pd
    g = globals()
    _ref = None
    # Prefer the 'sel' chosen in 总览
    if "sel" in g:
        try:
            _ref = pd.to_datetime(g["sel"], errors="coerce")
        except Exception:
            _ref = None
    # Fallback to 'sel2' (下钻)
    if (_ref is None) or (pd.isna(_ref)):
        if "sel2" in g:
            try:
                _ref = pd.to_datetime(g["sel2"], errors="coerce")
            except Exception:
                _ref = None
    # Final fallback: newest date in overview
    if (_ref is None) or (pd.isna(_ref)):
        try:
            _ref = pd.to_datetime(overview["日期"], errors="coerce").max()
        except Exception:
            _ref = None
    if _ref is None or pd.isna(_ref):
        return None
    return _ref.to_period("M")


# ===== 常量（同 v5） =====
_SALES_SHEETS    = ["销量-5001","销量-5002","销量"]
_TRANSFER_SHEETS = ["销量-转调理品原料","销量-调拨宫产量","销量-生肉转调理品原料"]
_DATE_CANDS   = ["单据日期","记帐日期","记账日期","凭证日期","日期","输入日期","过账日期"]
_CODE_CANDS   = ["物料","物料号","物料编码","物料编号","物料代码","Material"]
_REV_CANDS    = ["收入-折让 CNY","净收入 CNY","项目货币净值 CNY","收入PN00*n CNY","收入-折让"]
_NETW_CANDS   = ["净重","净重 KG","净重KG","数量(kg)","数量","重量","重量(kg)"]
_EXTAMT_CANDS = ["本币中的外部金额","本位币金额","本位币中的外部金额"]
_QTY_CANDS2   = ["数量","数量(kg)","净重","净重KG","净重 KG","重量","重量(kg)"]
_ARRIVE_CANDS = ["棚前-到场","棚前到场","到场","到场重量","到场(kg)","到场KG"]

SPECIFIED = [
    "腿类","胸类","胸类-胸","胸类-胸皮","里肌类","翅类","整鸡类","骨架类",
    "爪类","鸡肝类","鸡心类","脖类","鸡胗类","鸡头类","油类"
]
ORDER = SPECIFIED + ["下料类","其他内脏","鸡头+鸡脖+骨架","总计"]
BASE_FOR_TOTAL = [x for x in ORDER if x not in ("胸类","鸡头+鸡脖+骨架","总计")]
MAIN_PRODUCTS = {"腿类", "胸类", "胸类-胸", "胸类-胸皮", "里肌类", "翅类", "爪类"}


def _normalize_project_set(projects):
    present = set()
    for p in projects:
        if p is None or pd.isna(p):
            continue
        s = str(p).strip()
        if s:
            present.add(s)
    return present


def _dynamic_base_projects(projects):
    dynamic_base = set(BASE_FOR_TOTAL)
    present = _normalize_project_set(projects)
    if "胸类" in present:
        dynamic_base.discard("胸类-胸")
        dynamic_base.discard("胸类-胸皮")
        dynamic_base.add("胸类")
    else:
        dynamic_base.discard("胸类")
    return dynamic_base


def _dynamic_main_projects(projects):
    dynamic_main = set(MAIN_PRODUCTS)
    present = _normalize_project_set(projects)
    if "胸类" in present:
        dynamic_main.discard("胸类-胸")
        dynamic_main.discard("胸类-胸皮")
        dynamic_main.add("胸类")
    else:
        dynamic_main.discard("胸类")
    return dynamic_main


def _segment_metrics(seg, project_set, lw_sum):
    if seg is None or seg.empty or not project_set:
        return np.nan, np.nan, np.nan, np.nan
    scope = seg[seg["项目"].isin(project_set)]
    if scope.empty:
        return np.nan, np.nan, np.nan, np.nan
    qty = pd.to_numeric(scope["产量(kg)"], errors="coerce").sum(min_count=1)
    amt = pd.to_numeric(scope["含税金额"], errors="coerce").sum(min_count=1)
    rate = (qty / lw_sum * 100.0) if (pd.notna(lw_sum) and lw_sum > 0 and pd.notna(qty)) else np.nan
    unit = (amt / qty) if (pd.notna(qty) and qty != 0 and pd.notna(amt)) else np.nan
    return qty, amt, rate, unit


def _component_value_unit(total_qty, comp_amt):
    if not (pd.notna(total_qty) and total_qty != 0 and pd.notna(comp_amt)):
        return np.nan
    return comp_amt / total_qty


def _side_component_value_unit(total_unit, main_unit, total_qty, side_amt):
    if pd.notna(total_unit) and pd.notna(main_unit):
        return total_unit - main_unit
    return _component_value_unit(total_qty, side_amt)

def _find_col(cols, cands):
    for c in cands:
        if c in cols: return c
    return None

def _pick_col(cols, cands):
    for c in cands:
        if c in cols: return c
    return None


def _is_code_like_series(s: pd.Series) -> float:
    """Return ratio of values that look like material codes (e.g. ABO0100493658)."""
    if s is None:
        return 0.0
    try:
        t = s.astype(str).str.strip()
    except Exception:
        return 0.0
    t = t[(t != "") & t.notna()]
    if t.empty:
        return 0.0
    # 常见编码形态：字母+数字，长度较长；避免把中文描述误判成编码
    m = t.str.match(r"^[A-Za-z]{2,}\d{6,}$", na=False)
    return float(m.mean()) if len(m) else 0.0


def _pick_code_col(df: pd.DataFrame):
    """Pick best code column among candidates by 'code-like' score; fallback to first match."""
    cols = [str(c).strip() for c in df.columns]
    cands = [c for c in _CODE_CANDS if c in cols]
    if not cands:
        return None
    best = None
    best_score = -1.0
    for c in cands:
        score = _is_code_like_series(df[c].head(5000))
        if score > best_score:
            best = c
            best_score = score
    return best if best is not None else cands[0]

def parse_datecol(series):
    return pd.to_datetime(series, errors="coerce").dt.normalize()

def normalize_code(codes):
    try:
        # 可迭代（list/tuple/set/np.array/pd.Series）
        iterator = list(codes)
    except TypeError:
        # 单个值
        iterator = [codes]

    out = []
    for c in iterator:
        if c is None:
            continue
        s = str(c).strip()
        if not s:
            continue
        # 去掉 Excel 导致的形如 '123456.0' 的小数点尾巴
        if '.' in s:
            s = s.split('.')[0]
        out.append(s)
    return out


# ===== 文件 / Sheet 工具 =====
class SheetBundle:
    """轻量封装：用 dict 存 Sheet，兼容旧的 sheet_names 访问方式。"""
    def __init__(self, sheets):
        self._sheets = sheets or {}
        self.sheet_names = list(self._sheets.keys())

    def get_sheet(self, name):
        return self._sheets.get(name)


def _get_sheet_names(xls):
    if xls is None:
        return []
    if isinstance(xls, SheetBundle):
        return list(xls.sheet_names)
    if isinstance(xls, dict):
        return list(xls.keys())
    names = getattr(xls, "sheet_names", None)
    return list(names) if names is not None else []


def _read_sheet_df(xls, sheet_name):
    """兼容 SheetBundle / ExcelFile / 路径，失败则返回 None。"""
    if xls is None:
        return None
    if isinstance(xls, SheetBundle):
        return xls.get_sheet(sheet_name)
    if isinstance(xls, dict):
        return xls.get(sheet_name)
    try:
        return pd.read_excel(xls, sheet_name)
    except Exception:
        try:
            return pd.read_excel(io.BytesIO(xls), sheet_name)
        except Exception:
            return None


def _load_raw(uploaded):
    """上传控件/Path/str → bytes；不存在时返回 None。"""
    if uploaded is None:
        return None
    if isinstance(uploaded, (str, Path)):
        path = Path(uploaded)
        return path.read_bytes() if path.exists() else None
    getter = getattr(uploaded, "getvalue", None)
    if callable(getter):
        try:
            return getter()
        except Exception:
            pass
    try:
        return uploaded.read()
    except Exception:
        return None


def _load_excel(uploaded):
    """返回 pd.ExcelFile 或 SheetBundle（若已是），失败返回 None。"""
    if uploaded is None:
        return None
    if isinstance(uploaded, SheetBundle):
        return uploaded
    try:
        return pd.ExcelFile(uploaded)
    except Exception:
        raw = _load_raw(uploaded)
        if raw is None:
            return None
        try:
            return pd.ExcelFile(io.BytesIO(raw))
        except Exception:
            return None


def _get_name(uploaded):
    if uploaded is None:
        return ""
    if isinstance(uploaded, (str, Path)):
        return str(uploaded)
    return getattr(uploaded, "name", "") or ""


# ===== 主数据加载（仅上传） =====
def build_main_sheet_bundle(prod_file, sale_file, trans_file, transfer_file):
    """
    将四类主数据文件合并成 SheetBundle：
    - 产量
    - 销量
    - 销量-转调理品原料（可选）
    - 销量-调拨宫产量（可选）
    若找不到指定 sheet，则取首个 sheet。
    """
    targets = [
        ("产量", prod_file, False),
        ("销量", sale_file, False),
        ("销量-转调理品原料", trans_file, True),   # 可选
        ("销量-调拨宫产量", transfer_file, True),  # 可选
    ]
    sheets = {}
    errors = []

    for name, upload, optional in targets:
        if upload is None or (isinstance(upload, list) and len(upload) == 0):
            if not optional:
                errors.append(f"缺少文件：{name}")
            continue
        if name == "销量" and isinstance(upload, list):
            sale_frames = []
            for up in upload:
                xls = _load_excel(up)
                if xls is None:
                    errors.append(f"{_get_name(up) or name} 读取失败")
                    continue
                names = _get_sheet_names(xls)
                if not names:
                    errors.append(f"{_get_name(up) or name} 未发现可用工作表")
                    continue
                target_sheet = name if name in names else names[0]
                df = _read_sheet_df(xls, target_sheet)
                if df is None:
                    errors.append(f"{_get_name(up) or name} → {target_sheet} 读取失败")
                    continue
                try:
                    df.columns = [str(c).strip() for c in df.columns]
                except Exception:
                    pass
                sale_frames.append(df)
            if sale_frames:
                sheets[name] = pd.concat(sale_frames, ignore_index=True)
            else:
                errors.append("未读取到可用的销量文件")
            continue
        xls = _load_excel(upload)
        if xls is None:
            errors.append(f"{_get_name(upload) or name} 读取失败")
            continue
        names = _get_sheet_names(xls)
        if not names:
            errors.append(f"{_get_name(upload) or name} 未发现可用工作表")
            continue
        target_sheet = name if name in names else names[0]
        df = _read_sheet_df(xls, target_sheet)
        if df is None:
            errors.append(f"{_get_name(upload) or name} → {target_sheet} 读取失败")
            continue
        try:
            df.columns = [str(c).strip() for c in df.columns]
        except Exception:
            pass
        sheets[name] = df

    bundle = SheetBundle(sheets) if sheets else None
    return bundle, errors


def _attach_rate_display(df, df_lw):
    """
    只负责在展示前补回“产成率%”列：
    产成率(%) = 该部位(项目)的 产量(kg) / 当日 毛鸡净重(kg) * 100
    仅改显示层，不改内部计算。
    """
    try:
        import pandas as pd
        import numpy as np
        if df is None or len(df)==0: 
            return df
        out = df.copy()
        if "日期" not in out.columns or "产量(kg)" not in out.columns:
            return out
        # 标准化日期
        out["日期"] = pd.to_datetime(out["日期"], errors="coerce").dt.normalize()
        lw = df_lw.copy() if df_lw is not None else None
        if isinstance(lw, pd.DataFrame) and not lw.empty:
            # 容错列名
            cand = ["毛鸡净重(kg)","毛鸡净重","净重","净重(kg)"]
            vcol = next((c for c in cand if c in lw.columns), None)
            dcol = next((c for c in ["日期","交鸡日期","记帐日期","记账日期","凭证日期","过账日期"] if c in lw.columns), None)
            if vcol and dcol:
                lw2 = lw[[dcol, vcol]].copy()
                lw2.columns = ["日期","_lw"]
                lw2["日期"] = pd.to_datetime(lw2["日期"], errors="coerce").dt.normalize()
                lw_day = lw2.groupby("日期", as_index=False)["_lw"].sum()
                out = out.merge(lw_day, on="日期", how="left")
                out["_den"] = out["_lw"].where(out["_lw"].notna() & (out["_lw"]>0), np.nan)
                out["产成率%"] = (out["产量(kg)"] / out["_den"]) * 100.0
                out.drop(columns=[c for c in ["_lw","_den"] if c in out.columns], inplace=True)
        return out
    except Exception:
        return df


def format_thousands(df, cols):
    """Return a copy with selected numeric columns formatted as integer strings with thousands separators."""
    if df is None or not hasattr(df, "empty") or df.empty:
        return pd.DataFrame()
    out = df.copy()
    for c in cols:
        if c not in out.columns:
            continue
        out[c] = pd.to_numeric(out[c], errors="coerce").apply(
            lambda v: "" if pd.isna(v) else f"{int(round(v)):,}"
        )
    return out


def format_two_decimals(df, cols):
    """Return a copy with selected numeric columns formatted to two decimals (string)."""
    if df is None or not hasattr(df, "empty") or df.empty:
        return pd.DataFrame()
    out = df.copy()
    for c in cols:
        if c not in out.columns:
            continue
        out[c] = pd.to_numeric(out[c], errors="coerce").apply(
            lambda v: "" if pd.isna(v) else f"{v:.2f}"
        )
    return out


def add_sale_rate(df):
    """Append 产销率列=销量/产量；产量<=0或比值<=0则为 NaN。"""
    if df is None or not hasattr(df, "empty") or df.empty:
        return pd.DataFrame()
    out = df.copy()
    if ("产量(kg)" not in out.columns) or ("销量(kg)" not in out.columns):
        return out
    qty = pd.to_numeric(out["产量(kg)"], errors="coerce")
    sale = pd.to_numeric(out["销量(kg)"], errors="coerce")
    rate = np.where((qty > 0) & (sale / qty > 0), sale / qty, np.nan)
    out["产销率"] = rate
    return out


# Excel 显示格式（仅影响显示，不改变单元格真实数值）
EXCEL_INT_COLS = {
    "产量(kg)", "销量(kg)", "含税金额", "原产量(kg)", "调整量(kg)", "调整后产量(kg)",
    "生肉产量(吨)", "销量(吨)", "宰鸡量(千只)",
}
EXCEL_TWO_DEC_COLS = {"含税单价", "月均价", "产值(元/kg)", "均重(kg/只)"}
EXCEL_PERCENT_INT_COLS = {"产销率", "产销率(%)"}
EXCEL_PERCENT_TWO_DEC_COLS = {"产成率%", "产成率(%)", "调整后产成率(%)"}


def _build_excel_col_format_map(columns):
    out = {}
    for c in columns:
        name = str(c).strip()
        if name in EXCEL_INT_COLS:
            out[name] = "#,##0"
        elif name in EXCEL_TWO_DEC_COLS or name.endswith("产值(元/kg)"):
            out[name] = "0.00"
        elif name in EXCEL_PERCENT_INT_COLS:
            # 注意：值已是“百分数口径”（如 25 表示 25%），不能用 0%
            out[name] = '0"%"'
        elif (name in EXCEL_PERCENT_TWO_DEC_COLS) or ("%" in name):
            # 注意：值已是“百分数口径”（如 25 表示 25%），不能用 0%
            out[name] = '0.00"%"'
    return out


def _apply_excel_formats_by_header(ws, header_row, data_start_row, data_end_row, col_format_map):
    if ws is None or data_end_row < data_start_row or not col_format_map:
        return
    header_index = {}
    for col_idx in range(1, ws.max_column + 1):
        hv = ws.cell(row=header_row, column=col_idx).value
        if hv is None:
            continue
        header_index[str(hv).strip()] = col_idx
    for col_name, fmt in col_format_map.items():
        col_idx = header_index.get(str(col_name).strip())
        if not col_idx:
            continue
        for r in range(data_start_row, data_end_row + 1):
            cell = ws.cell(row=r, column=col_idx)
            if cell.value is None or isinstance(cell.value, str):
                continue
            cell.number_format = fmt


def _apply_excel_format_by_col_idx(ws, col_start, col_end, data_start_row, data_end_row, fmt):
    if ws is None or data_end_row < data_start_row:
        return
    for c in range(col_start, col_end + 1):
        for r in range(data_start_row, data_end_row + 1):
            cell = ws.cell(row=r, column=c)
            if cell.value is None or isinstance(cell.value, str):
                continue
            cell.number_format = fmt



def _get_or_create_writer_sheet(writer, sheet_name):
    ws = writer.sheets.get(sheet_name)
    if ws is not None:
        return ws
    book = writer.book
    if (
        len(book.sheetnames) == 1
        and book.sheetnames[0] == "Sheet"
        and book.active.max_row == 1
        and book.active.max_column == 1
        and book.active["A1"].value is None
    ):
        ws = book.active
        ws.title = sheet_name
    elif sheet_name in book.sheetnames:
        ws = book[sheet_name]
    else:
        ws = book.create_sheet(title=sheet_name)
    writer.sheets[sheet_name] = ws
    return ws


def _write_core_summary_excel_section(writer, sheet_name, start_row, df_sec, title):
    ws = _get_or_create_writer_sheet(writer, sheet_name)
    title_row = start_row + 1
    header_top_row = start_row + 2
    header_bottom_row = start_row + 3
    data_start_row = start_row + 4

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(bold=True)
    title_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.cell(row=title_row, column=1, value=title).font = title_font

    header_cells = {
        (header_top_row, 1): "日期",
        (header_top_row, 2): "产成率(%)",
        (header_top_row, 5): "产值(元/kg)",
        (header_top_row, 8): "生肉产量\n(吨)",
        (header_top_row, 9): "销量\n(吨)",
        (header_top_row, 10): "产销率\n(%)",
        (header_top_row, 11): "宰鸡量\n(千只)",
        (header_top_row, 12): "均重\n(kg/只)",
        (header_bottom_row, 2): "主产品",
        (header_bottom_row, 3): "副产品",
        (header_bottom_row, 4): "总计",
        (header_bottom_row, 5): "主产品",
        (header_bottom_row, 6): "副产品",
        (header_bottom_row, 7): "总计",
    }
    for (row_idx, col_idx), value in header_cells.items():
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    merges = [
        (header_top_row, 1, header_bottom_row, 1),
        (header_top_row, 2, header_top_row, 4),
        (header_top_row, 5, header_top_row, 7),
        (header_top_row, 8, header_bottom_row, 8),
        (header_top_row, 9, header_bottom_row, 9),
        (header_top_row, 10, header_bottom_row, 10),
        (header_top_row, 11, header_bottom_row, 11),
        (header_top_row, 12, header_bottom_row, 12),
    ]
    for r1, c1, r2, c2 in merges:
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        merged_cell = ws.cell(row=r1, column=c1)
        merged_cell.font = header_font
        merged_cell.alignment = center
        merged_cell.border = border

    col_specs = [
        ("日期", 14, None),
        ("主产品产成率(%)", 10, '0.00"%"'),
        ("副产品产成率(%)", 10, '0.00"%"'),
        ("产成率(%)", 10, '0.00"%"'),
        ("主产品产值(元/kg)", 10, "0.00"),
        ("副产品产值(元/kg)", 10, "0.00"),
        ("产值(元/kg)", 10, "0.00"),
        ("生肉产量(吨)", 12, "0.00"),
        ("销量(吨)", 12, "0.00"),
        ("产销率(%)", 10, '0.00"%"'),
        ("宰鸡量(千只)", 12, "0.00"),
        ("均重(kg/只)", 12, "0.00"),
    ]

    ws.row_dimensions[header_top_row].height = 22
    ws.row_dimensions[header_bottom_row].height = 22
    for col_idx, (_, width, _) in enumerate(col_specs, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_offset, (_, row) in enumerate(df_sec.iterrows()):
        excel_row = data_start_row + row_offset
        for col_idx, (src_col, _, fmt) in enumerate(col_specs, start=1):
            value = row.get(src_col)
            if pd.isna(value):
                value = None
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.alignment = center
            cell.border = border
            if fmt and value is not None and not isinstance(value, str):
                cell.number_format = fmt

    return len(df_sec) + 4


@st.cache_resource(show_spinner=False)
def read_all_sheets(uploaded):
    if uploaded is None:
        return None
    if isinstance(uploaded, SheetBundle):
        return uploaded
    return _load_excel(uploaded)
# ===== 读取产量（按日按物料号） =====
def read_qty_per_code_per_day(xls):
    empty = pd.DataFrame(columns=["日期","物料号","产量(kg)"])
    if xls is None:
        return empty

    target_sheet = None
    sheet_names = _get_sheet_names(xls)
    if "产量" in sheet_names:
        target_sheet = "产量"
    else:
        for name in sheet_names:
            name_str = str(name)
            if any(key in name_str for key in ("产量", "生产", "产出")):
                target_sheet = name
                break

    if target_sheet is None:
        try:
            sheets_preview = ", ".join(map(str, sheet_names[:5]))
        except Exception:
            sheets_preview = ""
        st.warning(f"主数据未找到名为“产量”的工作表（当前工作表：{sheets_preview}...），请确认上传的文件。")
        return empty

    df = _read_sheet_df(xls, target_sheet)
    if df is None:
        return empty
    df.columns = [str(c).strip() for c in df.columns]
    dcol = _pick_col(df.columns, ["日期","记帐日期","记账日期","凭证日期","过账日期","单据日期"])
    ccol = _pick_code_col(df)
    qcol = _pick_col(df.columns, ["数量(kg)","数量","净重","净重KG","净重 KG","重量","重量(kg)","KG","kg"])
    if not (dcol and ccol and qcol):
        st.warning("“产量”工作表缺少“日期/物料/数量(kg)”列，无法生成总览。")
        return empty
    out = pd.DataFrame({
        "日期": pd.to_datetime(df[dcol], errors="coerce").dt.normalize(),
        "物料号": df[ccol].astype(str).str.strip(),
        "产量(kg)": pd.to_numeric(df[qcol], errors="coerce")
    }).dropna(subset=["日期","物料号","产量(kg)"])
    return out.groupby(["日期","物料号"], as_index=False)["产量(kg)"].sum()

# ===== BOM 映射（含胸类细分与别名宽匹配） =====
@st.cache_data(show_spinner=False)
def read_bom_mapping(uploaded):
    code2major = {}
    code2desc = {}
    if uploaded is None: 
        return {}, {}
    xls = _load_excel(uploaded)
    if xls is None: 
        return {}, {}
    def canon_major(text: str) -> str:
        if not isinstance(text, str): return ""
        t = text.strip().replace(" ", "")
        if "胸" in t: return "胸类"
        if "腿" in t: return "腿类"
        if ("里肌" in t) or ("里脊" in t) or ("里脇" in t): return "里肌类"
        if "翅" in t: return "翅类"
        if "整鸡" in t: return "整鸡类"
        if "骨架" in t: return "骨架类"
        if ("鸡爪" in t) or (t.endswith("爪")) or ("爪" in t): return "爪类"
        if ("鸡肝" in t) or ("肝" in t and "鸡" in t): return "鸡肝类"
        if ("鸡心" in t) or ("心" in t and "鸡" in t): return "鸡心类"
        if ("脖" in t) or ("鸡脖" in t): return "脖类"
        if ("鸡头" in t) or ("头" in t and "鸡" in t): return "鸡头类"
        if ("鸡胗" in t) or ("鸡肫" in t) or (t in ["胗","胗类"]) or ("胗" in t): return "鸡胗类"
        if "油" in t: return "油类"
        if "下料" in t: return "下料类"
        if "内脏" in t: return "其他内脏"
        return t
    def _parse(df):
        cols = [str(c).strip() for c in df.columns]
        code_cands = [c for c in cols if any(k in c for k in ["物料号","物料编码","物料编号","物料代码","编码","代码","Material","物料"])]
        code_col = None
        if code_cands:
            # 同名语义列并存时，优先选择“更像编码”的列
            best_score = -1.0
            for c in code_cands:
                score = _is_code_like_series(df[c].head(5000))
                if score > best_score:
                    code_col = c
                    best_score = score
        maj_col  = next((c for c in cols if any(k in c for k in ["外部物料组描述","物料组描述","部位","部位大类","类别","大类","分类"])), None)
        sub_col  = next((c for c in cols if any(k in c for k in ["子类","二级","小类","品类","部位描述"])), None)
        desc_col = next((c for c in cols if (c not in {maj_col, sub_col}) and any(k in c for k in ["物料描述","长文本","描述","物料名称","品名","名称"])), None)
        if (code_col is None) or (maj_col is None): return
        use_cols = [code_col, maj_col]
        if sub_col: use_cols.append(sub_col)
        if desc_col: use_cols.append(desc_col)
        tmp = df[use_cols].dropna(subset=[code_col, maj_col]).copy()
        tmp[code_col] = normalize_code(tmp[code_col])
        tmp[maj_col]  = tmp[maj_col].astype(str).map(canon_major)
        for _, r in tmp.iterrows():
            code = str(r[code_col]).strip(); maj = str(r[maj_col]).strip()
            if not code or not maj: continue
            final_major = "胸类-胸皮" if (maj=="胸类" and ("胸皮" in str(r.get(sub_col,"")).replace(" ", ""))) else ("胸类-胸" if maj=="胸类" else maj)
            code2major[code] = final_major
            if desc_col:
                desc = str(r.get(desc_col, "")).strip()
                if desc:
                    code2desc[code] = desc
    try:
        for s in _get_sheet_names(xls):
            df = _read_sheet_df(xls, s)
            if df is not None:
                _parse(df)
    except Exception:
        pass
    # 强制归类（你的固定清单）
    def _force_put(codes, major):
        for v in normalize_code(codes):
            code2major[v] = major
    _force_put(["ABF0100322058","ABF0100493650","ABF0600493638","ADF0100320110","ADF0100323640","ADF0600493600"], "油类")
    _force_put(["ABO0900945618","ABZ0600493658"], "下料类")
    # 胸类-胸皮 强制映射
    _force_put(["ABB0700210118","ADB0700212840","ABB0700380618","ABB0700493610","ABB0700493638","ABB0900493611",
        "ADB0700000008","ADB0700000018","ADB0700000048","ADB0700000058","ADB0700380610","ADB0900945610"
    ], "胸类-胸皮")
    _force_put(["ABO0100493640","ABO0100493658","ABO0300210118","ABO0300211658","ABO0300493658",
                "ABO0800493648","ABZ0400723608","ADO0100210100","ADO0100380600","ADO0300210100",
                "ADO0300380600","ADO1500945610"], "其他内脏")
    return code2major, code2desc

# ===== 构建“原始”当日代码单价（销量 & 转调理/调拨宫） =====

def build_daily_code_price_raw(xls):
    sheets = _get_sheet_names(xls)
    if not sheets:
        return pd.DataFrame(columns=["日期","物料号","综合单价","金额","数量"])
    # --- 分别采集 销售 与 转调/调拨 两类来源 ---
    sale_frames = []
    trans_frames = []

    for sh in sheets:
        df = _read_sheet_df(xls, sh)
        if df is None:
            continue
        df.columns = [str(c).strip() for c in df.columns]

        # 公共列定位
        dcol = _pick_col(df.columns, _DATE_CANDS)
        ccol = _pick_code_col(df)
        if not (dcol and ccol):
            continue

        if sh in _SALES_SHEETS:
            # SaTy 为 zcr/zdr 的行不参与后续计算
            saty_col = next((c for c in df.columns if str(c).strip().lower() == "saty"), None)
            if saty_col:
                mask_exclude = df[saty_col].astype(str).str.strip().str.lower().isin({"zcr", "zdr"})
                if mask_exclude.any():
                    df = df.loc[~mask_exclude].copy()
            rev = _pick_col(df.columns, _REV_CANDS)   # 收入-折让 CNY（或同义列）
            wt  = _pick_col(df.columns, _NETW_CANDS)  # 净重/数量(kg)
            if not (rev and wt): 
                continue
            weight = pd.to_numeric(df[wt], errors="coerce")
            tmp = pd.DataFrame({
                "日期": pd.to_datetime(df[dcol], errors="coerce").dt.normalize(),
                "物料号": df[ccol].astype(str).str.strip(),
                "销售金额": pd.to_numeric(df[rev], errors="coerce"),
                "销售净重": weight
            }).dropna(subset=["日期","物料号"])
            if not tmp.empty:
                sale_frames.append(tmp.groupby(["日期","物料号"], as_index=False)[["销售金额","销售净重"]].sum())

        elif sh in _TRANSFER_SHEETS:
            ext = _pick_col(df.columns, _EXTAMT_CANDS)  # 本币中的外部金额/本位币金额
            qty = _pick_col(df.columns, _QTY_CANDS2)    # 数量(kg)/净重
            if not (ext and qty): 
                continue
            tmp = pd.DataFrame({
                "日期": pd.to_datetime(df[dcol], errors="coerce").dt.normalize(),
                "物料号": df[ccol].astype(str).str.strip(),
                "转调金额_含税": pd.to_numeric(df[ext], errors="coerce") * 1.09,
                "转调数量": pd.to_numeric(df[qty], errors="coerce")
            }).dropna(subset=["日期","物料号"])
            if not tmp.empty:
                trans_frames.append(tmp.groupby(["日期","物料号"], as_index=False)[["转调金额_含税","转调数量"]].sum())

    # 合并销售与转调/调拨
    if not sale_frames and not trans_frames:
        return pd.DataFrame(columns=["日期","物料号","综合单价","金额","数量"])

    sale  = pd.concat(sale_frames,  ignore_index=True) if sale_frames  else pd.DataFrame(columns=["日期","物料号","销售金额","销售净重"])
    trans = pd.concat(trans_frames, ignore_index=True) if trans_frames else pd.DataFrame(columns=["日期","物料号","转调金额_含税","转调数量"])

    comb = sale.merge(trans, on=["日期","物料号"], how="outer")

    # 统一优先级：能加则加；两边都空则 NaN
    has_any = comb[["销售金额","销售净重","转调金额_含税","转调数量"]].notna().any(axis=1)

    # 转调金额/数量改为使用绝对值参与合计
    trans_amt_abs = comb["转调金额_含税"].abs()
    trans_qty_abs = comb["转调数量"].abs()

    comb["金额"] = np.where(
        has_any,
        comb["销售金额"].fillna(0) + trans_amt_abs.fillna(0),
        np.nan
    )
    comb["数量"] = np.where(
        has_any,
        comb["销售净重"].fillna(0) + trans_qty_abs.fillna(0),
        np.nan
    )

    comb = comb[["日期","物料号","金额","数量","销售金额","销售净重"]]
    comb = comb.groupby(["日期","物料号"], as_index=False)[["金额","数量","销售金额","销售净重"]].sum(min_count=1)
    # 对指定物料号的金额统一乘 1.09（数量不变）；销量口径金额同样乘以系数，便于负价回落时保持口径一致
    if 'tax_factor_for_code' in globals():
        comb["金额"] = comb["金额"] * comb["物料号"].apply(tax_factor_for_code)
        comb["销售金额"] = comb["销售金额"] * comb["物料号"].apply(tax_factor_for_code)

    # 数量允许为负；仅数量==0/NaN 时置 NaN 避免除零
    qty = comb["数量"].replace(0, np.nan)
    comb["综合单价"] = comb["金额"] / qty

    # 负价回落：若合并价为负，改用“销量表口径”的单价（销售金额/销售净重，不含转调）
    sale_qty = comb["销售净重"].replace(0, np.nan)
    comb["_销售单价_only"] = comb["销售金额"] / sale_qty
    neg_mask = comb["综合单价"] < 0
    comb.loc[neg_mask, "综合单价"] = comb.loc[neg_mask, "_销售单价_only"]

    return comb[["日期","物料号","金额","数量","综合单价"]]


def build_daily_total_qty(xls):
    """
    用于总计销量口径：销售净重按日汇总 + 转调数量按日汇总后取绝对值。
    返回列：日期, 总销量(kg)
    """
    import pandas as pd, numpy as np
    sheets = _get_sheet_names(xls)
    if not sheets:
        return pd.DataFrame(columns=["日期","总销量(kg)"])
    sale_frames = []
    trans_frames = []
    for sh in sheets:
        df = _read_sheet_df(xls, sh)
        if df is None:
            continue
        df.columns = [str(c).strip() for c in df.columns]
        dcol = _pick_col(df.columns, _DATE_CANDS)
        if not dcol:
            continue
        if sh in _SALES_SHEETS:
            wt  = _pick_col(df.columns, _NETW_CANDS)
            if not wt:
                continue
            saty_col = next((c for c in df.columns if str(c).strip().lower() == "saty"), None)
            if saty_col:
                mask_exclude = df[saty_col].astype(str).str.strip().str.lower().isin({"zcr", "zdr"})
                if mask_exclude.any():
                    df = df.loc[~mask_exclude].copy()
            tmp = pd.DataFrame({
                "日期": pd.to_datetime(df[dcol], errors="coerce").dt.normalize(),
                "销售净重": pd.to_numeric(df[wt], errors="coerce")
            }).dropna(subset=["日期"])
            if not tmp.empty:
                sale_frames.append(tmp.groupby(["日期"], as_index=False)["销售净重"].sum())
        elif sh in _TRANSFER_SHEETS:
            qty = _pick_col(df.columns, _QTY_CANDS2)
            if not qty:
                continue
            tmp = pd.DataFrame({
                "日期": pd.to_datetime(df[dcol], errors="coerce").dt.normalize(),
                "转调数量": pd.to_numeric(df[qty], errors="coerce")
            }).dropna(subset=["日期"])
            if not tmp.empty:
                trans_frames.append(tmp.groupby(["日期"], as_index=False)["转调数量"].sum())
    sale = pd.concat(sale_frames, ignore_index=True) if sale_frames else pd.DataFrame(columns=["日期","销售净重"])
    trans = pd.concat(trans_frames, ignore_index=True) if trans_frames else pd.DataFrame(columns=["日期","转调数量"])
    out = sale.merge(trans, on="日期", how="outer")
    out["销售净重"] = pd.to_numeric(out["销售净重"], errors="coerce")
    out["转调数量"] = pd.to_numeric(out["转调数量"], errors="coerce")
    out["总销量(kg)"] = out["销售净重"].fillna(0) + out["转调数量"].abs().fillna(0)
    return out[["日期","总销量(kg)"]]


def fill_price_code_month_avg(pr_raw, qty, code2major, manual_month=None):
    """

    """
    import pandas as _pd
    import numpy as _np

    if pr_raw is None or pr_raw.empty or qty is None or qty.empty:
        return _pd.DataFrame(columns=["日期","物料号","综合单价_filled"])

    # 标准化
    pr = pr_raw.copy()
    pr["日期"] = _pd.to_datetime(pr["日期"], errors="coerce").dt.normalize()
    qty2 = qty.copy()
    qty2["日期"] = _pd.to_datetime(qty2["日期"], errors="coerce").dt.normalize()

    # 需要产出价格的 (日期, 物料号)
    need = qty2[["日期","物料号"]].drop_duplicates()
    need["月"] = need["日期"].dt.to_period("M")

    # 当日已有价（综合单价）
    pr["月"] = pr["日期"].dt.to_period("M")
    pr_day = pr[["日期","物料号","月","综合单价","金额","数量"]].copy()

    # 先左连当日实价
    out0 = need.merge(pr_day[["日期","物料号","月","综合单价"]], on=["日期","物料号","月"], how="left")

    # 已取消前瞻定价：缺价不再看 D+1/D+2，直接走历史加权，若仍无价则保持缺价

    # === 本月月均价回填（MTD）：月初到“当日”累计金额/累计数量 ===
    daily = (pr_day.groupby(["物料号", "月", "日期"], as_index=False)
                  .agg(金额=("金额", "sum"), 数量=("数量", "sum")))
    daily["金额"] = _pd.to_numeric(daily["金额"], errors="coerce").fillna(0.0)
    daily["数量"] = _pd.to_numeric(daily["数量"], errors="coerce").fillna(0.0)
    daily = daily.sort_values(["物料号", "月", "日期"])
    daily["金额累计"] = daily.groupby(["物料号", "月"])["金额"].cumsum()
    daily["数量累计"] = daily.groupby(["物料号", "月"])["数量"].cumsum()
    daily["综合单价_filled"] = _np.where(
        daily["数量累计"] != 0,
        daily["金额累计"] / daily["数量累计"],
        _np.nan
    )
    # 将“有价格日”的累计均价，向后匹配到该物料该月的每个需求日（<=当日最近一条）
    need_key = need[["日期", "物料号", "月"]].drop_duplicates().copy()
    fill_key = daily[["日期", "物料号", "月", "综合单价_filled"]].copy()
    if not fill_key.empty:
        parts = []
        for (code, mon), lk in need_key.groupby(["物料号", "月"], sort=False):
            rk = fill_key[(fill_key["物料号"] == code) & (fill_key["月"] == mon)].copy()
            lk = lk.sort_values("日期")
            rk = rk.sort_values("日期")
            if rk.empty:
                lk["综合单价_filled"] = _np.nan
                parts.append(lk)
                continue
            mk = _pd.merge_asof(
                lk,
                rk[["日期", "综合单价_filled"]],
                on="日期",
                direction="backward",
                allow_exact_matches=True,
            )
            parts.append(mk)
        code_filled = _pd.concat(parts, ignore_index=True) if parts else need_key.assign(综合单价_filled=_np.nan)
    else:
        code_filled = need_key.copy()
        code_filled["综合单价_filled"] = _np.nan

    out = out0.merge(code_filled, on=["日期", "物料号", "月"], how="left")
    out["综合单价_filled"] = out["综合单价"].where(out["综合单价"].notna(), out["综合单价_filled"])

    if manual_month is not None and not getattr(manual_month, 'empty', True):
        mm = manual_month.copy()
        mm['物料号'] = mm['物料号'].astype(str).str.strip()
        out = out.merge(mm.rename(columns={'手工单价':'_手工单价'}), on='物料号', how='left')
        out['综合单价_filled'] = out['综合单价_filled'].where(out['综合单价_filled'].notna(), out['_手工单价'])
        if '_手工单价' in out.columns:
            out.drop(columns=['_手工单价'], inplace=True)

    return out[["日期","物料号","综合单价_filled"]]


def build_overview(xls, code2major, df_lw=None, manual_month_df=None, alloc_long=None):
    # 产量
    qty = read_qty_per_code_per_day(xls)
    if qty.empty: return pd.DataFrame(), pd.DataFrame()
    # 销售价原始
    pr_raw = build_daily_code_price_raw(xls)
    # 价格断档填充：本月月初至当日（MTD）平均价
    pr_fill = fill_price_code_month_avg(pr_raw, qty, code2major, manual_month_df)

    m = qty.merge(pr_fill, on=["日期","物料号"], how="left")

    # << 部位分摊：若命中分摊规则则按比例拆分，否则回落到一对一映射 >>
    m_split = apply_part_allocation(m, alloc_long, code2major)

    # 销量（按项目拆分）
    sales_proj = pd.DataFrame(columns=["日期","项目","销量(kg)"])
    try:
        if pr_raw is not None and not pr_raw.empty:
            sales_base = pr_raw.copy()
            sales_base["日期"] = pd.to_datetime(sales_base["日期"], errors="coerce").dt.normalize()
            if "数量" not in sales_base.columns:
                sales_base["数量"] = np.nan
            if "综合单价" not in sales_base.columns:
                sales_base["综合单价"] = np.nan
            sales_base = sales_base.rename(columns={"数量":"产量(kg)","综合单价":"综合单价_filled"})
            sales_split = apply_part_allocation(sales_base, alloc_long, code2major)
            if sales_split is not None and not sales_split.empty:
                sales_proj = sales_split.groupby(["日期","项目"], as_index=False)["产量(kg)"].sum()
                sales_proj.rename(columns={"产量(kg)":"销量(kg)"}, inplace=True)
                frames = []
                for d, g in sales_proj.groupby("日期"):
                    g2 = g.copy(); add = []
                    chest = g2[g2["项目"].isin(["胸类-胸","胸类-胸皮"])]
                    if not chest.empty:
                        add.append({"日期": d, "项目": "胸类", "销量(kg)": chest["销量(kg)"].sum()})
                    comb = g2[g2["项目"].isin(["鸡头类","脖类","骨架类"])]
                    if not comb.empty:
                        add.append({"日期": d, "项目": "鸡头+鸡脖+骨架", "销量(kg)": comb["销量(kg)"].sum()})
                    if add:
                        g2 = pd.concat([g2, pd.DataFrame(add)], ignore_index=True)
                    frames.append(g2)
            sales_proj = pd.concat(frames, ignore_index=True) if frames else sales_proj
    except Exception:
        sales_proj = pd.DataFrame(columns=["日期","项目","销量(kg)"])

    over = m_split.groupby(["日期","项目"], as_index=False).agg({"产量(kg)":"sum","含税金额":"sum"})
    over["含税单价"] = np.where(over["产量(kg)"] != 0, over["含税金额"]/over["产量(kg)"], np.nan)
    over["项目"] = pd.Categorical(over["项目"], categories=ORDER, ordered=True)
    over = over.sort_values("项目")

    # 胸类（汇总）与组合行
    frames=[]
    for d,g in over.groupby("日期"):
        g2=g.copy(); add=[]
        chest=g2[g2["项目"].isin(["胸类-胸","胸类-胸皮"])]
        if not chest.empty:
            s=chest[["产量(kg)","含税金额"]].sum(); qtyc=s["产量(kg)"]; amtc=s["含税金额"]
            add.append({"日期":d,"项目":"胸类","产量(kg)":qtyc,"含税金额":amtc,"含税单价":(amtc/qtyc if qtyc>0 else 0)})
        comb=g2[g2["项目"].isin(["鸡头类","脖类","骨架类"])]
        if not comb.empty:
            s=comb[["产量(kg)","含税金额"]].sum(); q=s["产量(kg)"]; a=s["含税金额"]
            add.append({"日期":d,"项目":"鸡头+鸡脖+骨架","产量(kg)":q,"含税金额":a,"含税单价":(a/q if q>0 else 0)})
        if add:
            g2=pd.concat([g2,pd.DataFrame(add)],ignore_index=True)
        frames.append(g2)
    over=pd.concat(frames,ignore_index=True)

    if sales_proj is not None and not sales_proj.empty:
        over = over.merge(sales_proj, on=["日期","项目"], how="left")
    if "销量(kg)" not in over.columns:
        over["销量(kg)"] = np.nan

    # 子类（分摊后下钻）：按“部位大类 + 物料号”
    minors = m_split.rename(columns={"项目":"部位大类","物料号":"子类"})
    minors = minors.groupby(["日期","部位大类","子类"],as_index=False).agg({"产量(kg)":"sum","含税金额":"sum"})
    _num = pd.to_numeric(minors["含税金额"], errors="coerce")
    _den = pd.to_numeric(minors["产量(kg)"], errors="coerce")
    minors["含税单价"] = np.divide(_num, _den, out=np.full(len(minors), np.nan), where=_den!=0)
    return over, minors

# ===== 读取净重台账（与 v5 相同） =====
@st.cache_data(show_spinner=False)
def read_liveweight(uploaded):
    """
    读取净重台账，支持：
    - Excel/CSV
    - 日期为文本/日期/Excel 数值序列
    - 同时提取交鸡量列（若存在），便于均重计算
    输出列：日期、毛鸡净重(kg)，可选：交鸡量
    """
    empty = pd.DataFrame(columns=["日期","毛鸡净重(kg)"])
    if uploaded is None:
        return empty

    def _maybe_scale_ton(df, wcol):
        """
        若净重数据疑似以“吨”为单位（数值普遍<5000），自动乘以1000转为 kg。
        """
        try:
            series = pd.to_numeric(df[wcol], errors="coerce")
            if series.dropna().empty:
                return df
            max_v = series.max()
            med_v = series.median()
            if pd.notna(max_v) and pd.notna(med_v) and (max_v <= 5000) and (med_v > 10):
                df[wcol] = series * 1000.0
                st.info("已自动将净重数据从“吨”换算为 kg（检测到数值普遍小于5000）。")
            else:
                df[wcol] = series
        except Exception:
            pass
        return df

    def _safe_day(x):
        if x is None or (isinstance(x, float) and pd.isna(x)) or (isinstance(x, str) and x.strip()==""):
            return pd.NaT
        if isinstance(x, (int, float)) and not isinstance(x, bool):
            try:
                return (pd.Timestamp("1899-12-30") + pd.to_timedelta(float(x), unit="D")).normalize()
            except Exception:
                pass
        t = pd.to_datetime(x, errors="coerce")
        return t.normalize() if pd.notna(t) else pd.NaT

    raw = _load_raw(uploaded)
    name = _get_name(uploaded).lower()
    try:
        if name.endswith(".csv"):
            if raw is None:
                return empty
            df = pd.read_csv(io.BytesIO(raw))
            dcol = _find_col(df.columns, ["日期","交鸡日期","记帐日期","记账日期","凭证日期","过账日期"])
            wcol = _find_col(df.columns, ["毛鸡净重(kg)","毛鸡净重","净重","净重(kg)"])
            ccol = _find_col(df.columns, ["交鸡量","交鸡只数","交鸡数量","鸡只数","只数","数量(只)","数量（只）","数量"])
            if dcol and wcol:
                out = pd.DataFrame({
                    "日期": df[dcol].apply(_safe_day),
                    "毛鸡净重(kg)": pd.to_numeric(df[wcol], errors="coerce")
                })
                out = _maybe_scale_ton(out, "毛鸡净重(kg)")
                if ccol:
                    out["交鸡量"] = pd.to_numeric(df[ccol], errors="coerce")
                out = out[out["日期"].notna() & out["毛鸡净重(kg)"].notna()]
                if not out.empty:
                    agg_cols = {"毛鸡净重(kg)":"sum"}
                    if "交鸡量" in out.columns:
                        agg_cols["交鸡量"] = "sum"
                    return out.groupby("日期", as_index=False).agg(agg_cols)
        else:
            xls = _load_excel(uploaded)
            if xls is None:
                return empty
            parts = []
            for s in _get_sheet_names(xls):
                df = _read_sheet_df(xls, s)
                if df is None:
                    continue
                df.columns = [str(c).strip() for c in df.columns]
                dcol = _find_col(df.columns, ["日期","交鸡日期","记帐日期","记账日期","凭证日期","过账日期"])
                wcol = _find_col(df.columns, ["毛鸡净重(kg)","毛鸡净重","净重","净重(kg)"])
                ccol = _find_col(df.columns, ["交鸡量","交鸡只数","交鸡数量","鸡只数","只数","数量(只)","数量（只）","数量"])
                if dcol and wcol:
                    out = pd.DataFrame({
                        "日期": df[dcol].apply(_safe_day),
                        "毛鸡净重(kg)": pd.to_numeric(df[wcol], errors="coerce")
                    })
                    out = _maybe_scale_ton(out, "毛鸡净重(kg)")
                    if ccol:
                        out["交鸡量"] = pd.to_numeric(df[ccol], errors="coerce")
                    out = out[out["日期"].notna() & out["毛鸡净重(kg)"].notna()]
                    if not out.empty:
                        parts.append(out)
            if parts:
                full = pd.concat(parts, ignore_index=True)
                agg_cols = {"毛鸡净重(kg)":"sum"}
                if "交鸡量" in full.columns:
                    agg_cols["交鸡量"] = "sum"
                return full.groupby("日期", as_index=False).agg(agg_cols)
    except Exception:
        pass
    return empty


@st.cache_data(show_spinner=False)
def read_shed_arrival(uploaded):
    """
    读取“棚前-到场”表（Excel/CSV），输出列：日期、棚前-到场。
    """
    empty = pd.DataFrame(columns=["日期","棚前-到场"])
    if uploaded is None:
        return empty

    def _safe_day(x):
        if x is None or (isinstance(x, float) and pd.isna(x)) or (isinstance(x, str) and x.strip()==""):
            return pd.NaT
        if isinstance(x, (int, float)) and not isinstance(x, bool):
            try:
                return (pd.Timestamp("1899-12-30") + pd.to_timedelta(float(x), unit="D")).normalize()
            except Exception:
                pass
        t = pd.to_datetime(x, errors="coerce")
        return t.normalize() if pd.notna(t) else pd.NaT

    raw = _load_raw(uploaded)
    name = _get_name(uploaded).lower()
    try:
        if name.endswith(".csv"):
            if raw is None:
                return empty
            df = pd.read_csv(io.BytesIO(raw))
            df.columns = [str(c).strip() for c in df.columns]
            dcol = _find_col(df.columns, _DATE_CANDS)
            vcol = _find_col(df.columns, _ARRIVE_CANDS)
            if dcol and vcol:
                out = pd.DataFrame({
                    "日期": df[dcol].apply(_safe_day),
                    "棚前-到场": pd.to_numeric(df[vcol], errors="coerce")
                })
                out = out.dropna(subset=["日期","棚前-到场"])
                if not out.empty:
                    return out.groupby("日期", as_index=False)["棚前-到场"].sum()
        else:
            xls = _load_excel(uploaded)
            if xls is None:
                return empty
            parts = []
            for s in _get_sheet_names(xls):
                df = _read_sheet_df(xls, s)
                if df is None:
                    continue
                df.columns = [str(c).strip() for c in df.columns]
                dcol = _find_col(df.columns, _DATE_CANDS)
                vcol = _find_col(df.columns, _ARRIVE_CANDS)
                if dcol and vcol:
                    out = pd.DataFrame({
                        "日期": df[dcol].apply(_safe_day),
                        "棚前-到场": pd.to_numeric(df[vcol], errors="coerce")
                    })
                    out = out.dropna(subset=["日期","棚前-到场"])
                    if not out.empty:
                        parts.append(out)
            if parts:
                full = pd.concat(parts, ignore_index=True)
                return full.groupby("日期", as_index=False)["棚前-到场"].sum()
    except Exception:
        pass
    return empty


def merge_liveweight_with_arrival(df_lw, df_arrival):
    """
    将“棚前-到场”按日加总后从净重中扣除，返回合并后的净重表。
    """
    try:
        import pandas as pd
        base_empty = (df_lw is None) or (not isinstance(df_lw, pd.DataFrame)) or df_lw.empty
        if df_arrival is None or not isinstance(df_arrival, pd.DataFrame) or df_arrival.empty:
            return df_lw

        base = df_lw.copy() if not base_empty else pd.DataFrame()
        wcol = next((c for c in ["毛鸡净重(kg)","毛鸡净重","净重","净重(kg)"] if c in base.columns), None)
        dcol = next((c for c in _DATE_CANDS if c in base.columns), None)
        ccol = next((c for c in ["交鸡量","交鸡只数","交鸡数量","鸡只数","只数","数量(只)","数量（只）","数量"] if c in base.columns), None)
        if dcol is None:
            dcol = "日期"
            base[dcol] = pd.NaT
        if wcol is None:
            wcol = "毛鸡净重(kg)"
            base[wcol] = np.nan
        cols = [dcol, wcol] + ([ccol] if ccol else [])
        base = base[cols]
        base[dcol] = pd.to_datetime(base[dcol], errors="coerce").dt.normalize()
        base[wcol] = pd.to_numeric(base[wcol], errors="coerce")
        if ccol:
            base[ccol] = pd.to_numeric(base[ccol], errors="coerce")
        base = base[base[dcol].notna() & base[wcol].notna()]
        agg = {wcol: "sum"}
        if ccol:
            agg[ccol] = "sum"
        base = base.groupby(dcol, as_index=False).agg(agg)

        arrival = df_arrival.copy()
        if "日期" not in arrival.columns or "棚前-到场" not in arrival.columns:
            return df_lw if not base_empty else base
        arrival["日期"] = pd.to_datetime(arrival["日期"], errors="coerce").dt.normalize()
        arrival["棚前-到场"] = pd.to_numeric(arrival["棚前-到场"], errors="coerce")
        arrival = arrival.dropna(subset=["日期","棚前-到场"])
        if arrival.empty:
            return df_lw if not base_empty else base
        arrival = arrival.groupby("日期", as_index=False)["棚前-到场"].sum()

        merged = base.merge(arrival, on="日期", how="outer")
        merged[wcol] = merged[wcol].fillna(0) - merged["棚前-到场"].fillna(0)
        merged.drop(columns=["棚前-到场"], inplace=True, errors="ignore")
        return merged
    except Exception:
        return df_lw


# ===== 读取补价表（按月，物料号→含税单价） =====
@st.cache_data(show_spinner=False)
def read_manual_month_price(uploaded):
    """
    读取手工补价表（Excel 或 CSV）。支持列名同义：
      - 物料列：物料号/物料编码/物料
      - 价格列：含税单价/综合单价/单价/售价/价格/PRICE
    """
    import pandas as _pd
    if uploaded is None:
        return _pd.DataFrame(columns=['物料号','手工单价'])
    try:
        raw = _load_raw(uploaded)
        name = _get_name(uploaded).lower()
        if raw is None:
            return _pd.DataFrame(columns=['物料号','手工单价'])
        if name.endswith(".csv"):
            df = _pd.read_csv(io.BytesIO(raw))
        else:
            df = _pd.read_excel(io.BytesIO(raw))
        df.columns = [str(c).strip() for c in df.columns]
        ccol = next((c for c in ['物料号','物料编码','物料','Material'] if c in df.columns), None)
        pcol = next((c for c in ['含税单价','综合单价','单价','售价','价格','PRICE','price'] if c in df.columns), None)
        if not (ccol and pcol):
            return _pd.DataFrame(columns=['物料号','手工单价'])
        out = _pd.DataFrame({
            '物料号': df[ccol].astype(str).str.strip(),
            '手工单价': _pd.to_numeric(df[pcol], errors='coerce')
        })
        out = out.dropna(subset=['物料号','手工单价'])
        out = out[out['物料号']!='']
        out = out.groupby('物料号', as_index=False)['手工单价'].mean()
        return out
    except Exception:
        return _pd.DataFrame(columns=['物料号','手工单价'])
# ===== 读取“部位分摊表”并标准化为长表 =====

@st.cache_data(show_spinner=False)
def read_part_allocation(uploaded):
    """
    输入：分摊表（Excel/CSV）。推荐列：日期(可选)、物料号(必填)、部位列(若干)、合计(可选)；部位列可为百分比或0~1。
    输出：长表：['日期','物料号','项目','权重']；其中“项目”为模型的大类名。
    匹配优先级：按日(日期+物料号) > 通用(仅物料号)。
    规则：填写“日期”→仅该日生效；未填写→默认通用（当月所有日期生效）。
    """
    import pandas as pd, io as _io

    if uploaded is None:
        return pd.DataFrame(columns=["日期","物料号","项目","权重"])

    raw = _load_raw(uploaded)
    name = _get_name(uploaded).lower()
    if raw is None:
        return pd.DataFrame(columns=["日期","物料号","项目","权重"])
    try:
        if name.endswith(".csv"):
            df = pd.read_csv(_io.BytesIO(raw))
        else:
            df = pd.read_excel(_io.BytesIO(raw), sheet_name=0)
    except Exception:
        return pd.DataFrame(columns=["日期","物料号","项目","权重"])

    df.columns = [str(c).strip() for c in df.columns]

    # 键列
    dcol = next((c for c in ["日期","单据日期","记帐日期","记账日期","凭证日期","过账日期"] if c in df.columns), None)
    ccol = next((c for c in ["物料号","物料","物料编码","物料编号","物料代码","Material"] if c in df.columns), None)
    if not ccol:
        return pd.DataFrame(columns=["日期","物料号","项目","权重"])
    rcol = next((c for c in ["比例","权重","占比","比重","ratio","Ratio","RATE"] if c in df.columns), None)

    # 识别分摊列（排除键列/描述列）
    exclude = {dcol, ccol, rcol, "公司","工厂","物料描述","品类","序号","部位","合计", None}
    part_cols = [c for c in df.columns if c not in exclude]

    # 列名映射到模型大类（含“鸡骨架/腿骨泥/半架”统一到 骨架类；“下料/地脚料”统一到 下料类）
    part2major = {
        "全腿":"腿类", "腿":"腿类",
        "胸肉块":"胸类-胸","胸":"胸类-胸","胸类-胸":"胸类-胸","胸类-胸皮":"胸类-胸皮",
        "带筋里":"里肌类","里肌":"里肌类",
        "全翅":"翅类","翅":"翅类",
        "爪":"爪类","鸡爪":"爪类",
        "鸡肝":"鸡肝类","肝":"鸡肝类",
        "鸡心":"鸡心类","心":"鸡心类",
        "去皮脖":"脖类","脖":"脖类","鸡脖":"脖类",
        "鸡头":"鸡头类","头":"鸡头类",
        "半架":"骨架类","鸡骨架":"骨架类","腿骨泥":"骨架类","腿骨":"骨架类","骨架":"骨架类",
        "胃":"鸡胗类","鸡胗":"鸡胗类","胗":"鸡胗类",
        "下料":"下料类","地脚料":"下料类","下料类":"下料类",
        "其他内脏":"其他内脏"
    }

    keep_pairs = []
    for c in part_cols:
        std = part2major.get(c, c)  # 未在词典中的，按原名尝试（若已是模型标准名会保留）
        keep_pairs.append((c, std))
    if not keep_pairs:
        return pd.DataFrame(columns=["日期","物料号","项目","权重"])

    # —— 关键：安全解析日期（含 Excel 数值日期），避免对 NaT 调 .normalize() ——
    def _safe_day(x):
        if x is None or (isinstance(x, float) and pd.isna(x)) or (isinstance(x, str) and x.strip()==""):
            return pd.NaT
        if isinstance(x, (int, float)) and not isinstance(x, bool):
            # Excel 序列日期：1899-12-30 起算
            try:
                return (pd.Timestamp("1899-12-30") + pd.to_timedelta(float(x), unit="D")).normalize()
            except Exception:
                pass
        t = pd.to_datetime(x, errors="coerce")
        return t.normalize() if pd.notna(t) else pd.NaT

    rows = []
    bad_rows = []
    for _, r in df.iterrows():
        code = str(r[ccol]).strip() if ccol in df.columns else ""
        if not code or code.lower()=="nan":
            continue
        day = _safe_day(r.get(dcol, None)) if dcol else pd.NaT

        # 行比例（可选）：用于“部分产量走不同方案”
        ratio = 1.0
        if rcol and rcol in df.columns:
            try:
                ratio = float(r.get(rcol, 1.0))
                if ratio > 1.0 and ratio <= 100.0:
                    ratio = ratio / 100.0
            except Exception:
                ratio = 1.0
        if ratio <= 0:
            continue

        weights = []
        for raw_name, std_name in keep_pairs:
            val = r.get(raw_name, None)
            if val is None or (isinstance(val, float) and pd.isna(val)):
                continue
            s = str(val).strip()
            if not s:
                continue
            try:
                v = float(s.replace("%",""))
                if ("%" in s) or v>1.001:  # 兼容百分比/100制
                    v = v/100.0
            except Exception:
                continue
            if v>0:
                weights.append((std_name, v))

        if not weights:
            continue
        tot = sum(v for _, v in weights)
        if tot<=0:
            continue
        # 严格校验：行内合计必须为 1（允许极小误差）
        if abs(tot - 1.0) > 1e-6:
            bad_rows.append({
                "物料号": code,
                "日期": day,
                "合计": tot,
            })
            continue

        for name, v in weights:
            rows.append({"日期": day, "物料号": code, "项目": name, "权重": float(v) * ratio})

    out = pd.DataFrame(rows, columns=["日期","物料号","项目","权重"])
    if not out.empty:
        out["日期"] = pd.to_datetime(out["日期"], errors="coerce")  # 不再 .dt.normalize()，已在 _safe_day 处理
    if bad_rows:
        try:
            st.error(f"部位分摊表存在 {len(bad_rows)} 行部位比例合计≠1，已忽略这些行。")
            st.dataframe(pd.DataFrame(bad_rows), use_container_width=True)
        except Exception:
            pass
    return out


@st.cache_data(show_spinner=False)
def read_restore_config(_xls):
    import pandas as _pd
    if _xls is None:
        return _pd.DataFrame(columns=["需要还原的部位","品项","原部位","权重值","原部位_标准"])
    try:
        if "部位还原配置" not in _get_sheet_names(_xls):
            return _pd.DataFrame(columns=["需要还原的部位","品项","原部位","权重值","原部位_标准"])
        df = _read_sheet_df(_xls, "部位还原配置")
        if df is None:
            return _pd.DataFrame(columns=["需要还原的部位","品项","原部位","权重值","原部位_标准"])
        df.columns = [str(c).strip() for c in df.columns]
        # 允许列名的宽松匹配
        col_map = {}
        for k in ["需要还原的部位","品项","原部位","权重值","原部位_标准"]:
            for c in df.columns:
                if str(c).strip() == k:
                    col_map[k] = c
                    break
        # 必须列
        need = ["需要还原的部位","品项","原部位","权重值"]
        if not all(k in col_map for k in need):
            return _pd.DataFrame(columns=["需要还原的部位","品项","原部位","权重值","原部位_标准"])
        out = _pd.DataFrame({
            "需要还原的部位": df[col_map["需要还原的部位"]].astype(str).str.strip(),
            "品项": df[col_map["品项"]].astype(str).str.strip(),
            "原部位": df[col_map["原部位"]].astype(str).str.strip(),
            "权重值": _pd.to_numeric(df[col_map["权重值"]], errors="coerce"),
        })
        if "原部位_标准" in col_map:
            out["原部位_标准"] = df[col_map["原部位_标准"]].astype(str).str.strip().where(df[col_map["原部位_标准"]].notna(), None)
        else:
            out["原部位_标准"] = None
        out = out.dropna(subset=["需要还原的部位","品项","原部位","权重值"])
        out = out[out["权重值"]!=0]
        # 别名：将“其中：胸部/胸皮”规范为胸类-胸/胸类-胸皮（若未提供标准列）
        alias = {"其中：胸部":"胸类-胸","其中：胸皮":"胸类-胸皮"}
        out["原部位_标准"] = out["原部位_标准"].where(out["原部位_标准"].notna() & (out["原部位_标准"]!=""), out["原部位"].map(alias))
        return out
    except Exception:
        return _pd.DataFrame(columns=["需要还原的部位","品项","原部位","权重值","原部位_标准"])




def read_restore_mapping_upload(uploaded):
    """Read uploaded restore mapping workbook/csv into code -> target parts."""
    mapping = {}
    if uploaded is None:
        return mapping
    try:
        frames = []
        name = _get_name(uploaded).lower()
        if name.endswith(".csv"):
            raw = _load_raw(uploaded)
            if raw is None:
                return mapping
            frames = [pd.read_csv(io.BytesIO(raw))]
        else:
            xls = _load_excel(uploaded)
            if xls is None:
                return mapping
            for s in _get_sheet_names(xls):
                df = _read_sheet_df(xls, s)
                if df is not None:
                    frames.append(df)

        for df in frames:
            cols = [str(c).strip() for c in df.columns]
            code_col = next((c for c in cols if any(k in c for k in ["\u7269\u6599\u53f7", "\u7269\u6599\u7f16\u7801", "\u7269\u6599\u7f16\u53f7", "\u7269\u6599\u4ee3\u7801", "\u7f16\u7801", "Material", "\u54c1\u9879", "\u7269\u6599"])), None)
            tgt_col = next((c for c in cols if any(k in c for k in ["\u76ee\u6807", "\u8fd8\u539f", "\u53bb\u5411", "\u5206\u914d\u5230", "\u90e8\u4f4d"])), None)
            if not (code_col and tgt_col):
                continue
            subset = df[[code_col, tgt_col]].dropna(subset=[code_col, tgt_col]).copy()
            for _, row in subset.iterrows():
                codes = normalize_code([row[code_col]])
                if not codes:
                    continue
                code = codes[0]
                tgt_raw = str(row[tgt_col]).strip()
                if not tgt_raw:
                    continue
                for sep in ["\uFF0C", "\uFF1B", "/", "\u3001", "|", ";"]:
                    tgt_raw = tgt_raw.replace(sep, ",")
                pieces = [t.strip() for t in tgt_raw.split(",") if t.strip()]
                if not pieces:
                    continue
                existing = mapping.get(code, [])
                for part in pieces:
                    if part not in existing:
                        existing.append(part)
                mapping[code] = existing
        return mapping
    except Exception:
        return {}


def _unify_restore_part(name: str) -> str:
    alias_extra = {
        "\u817f": "\u817f\u7c7b",
        "\u80f8": "\u80f8\u7c7b",
        "\u80f8\u90e8": "\u80f8\u7c7b-\u80f8",
        "\u80f8\u76ae": "\u80f8\u7c7b-\u80f8\u76ae",
        "\u5176\u4e2d\uFF1A\u80f8\u90e8": "\u80f8\u7c7b-\u80f8",
        "\u5176\u4e2d\uFF1A\u80f8\u76ae": "\u80f8\u7c7b-\u80f8\u76ae",
        "\u91cc\u808c": "\u91cc\u808c\u7c7b",
        "\u7fc5": "\u7fc5\u7c7b",
        "\u9e21\u9aa8\u67b6": "\u9aa8\u67b6\u7c7b",
        "\u9aa8\u67b6": "\u9aa8\u67b6\u7c7b",
        "\u9e21\u5934": "\u9e21\u5934\u7c7b",
        "\u9e21\u809d": "\u9e21\u809d\u7c7b",
        "\u9e21\u5fc3": "\u9e21\u5fc3\u7c7b",
        "\u8116": "\u8116\u7c7b",
        "\u9e21\u8116": "\u8116\u7c7b",
        "\u6574\u9e21": "\u6574\u9e21\u7c7b",
        "\u5176\u4ed6\u5185\u810f": "\u5176\u4ed6\u5185\u810f",
        "\u4e0b\u6599": "\u4e0b\u6599\u7c7b",
        "\u4e0b\u6599\u7c7b": "\u4e0b\u6599\u7c7b",
    }
    x = str(name).strip()
    try:
        if 'ALIAS' in globals() and x in ALIAS:
            return ALIAS[x]
    except Exception:
        pass
    return alias_extra.get(x, x)


def _restore_lw_on(df_lw, day_value):
    val = float("nan")
    try:
        if df_lw is not None and not df_lw.empty:
            dcol = next((c for c in ["\u65e5\u671f", "\u4ea4\u9e21\u65e5\u671f", "\u8bb0\u5e10\u65e5\u671f", "\u8bb0\u8d26\u65e5\u671f", "\u51ed\u8bc1\u65e5\u671f", "\u8fc7\u8d26\u65e5\u671f"] if c in df_lw.columns), None)
            vcol = next((c for c in ["\u6bdb\u9e21\u51c0\u91cd(kg)", "\u6bdb\u9e21\u51c0\u91cd", "\u51c0\u91cd", "\u51c0\u91cd(kg)"] if c in df_lw.columns), None)
            if dcol and vcol:
                _lw = df_lw[[dcol, vcol]].copy()
                _lw.columns = ["\u65e5\u671f", "_lw"]
                _lw["\u65e5\u671f"] = pd.to_datetime(_lw["\u65e5\u671f"], errors="coerce").dt.normalize()
                val = _lw.loc[_lw["\u65e5\u671f"] == day_value, "_lw"].sum()
    except Exception:
        val = float("nan")
    return val


def _build_restore_rate_dict(ov_slice, lw_value):
    rate_df = ov_slice.groupby("\u9879\u76ee", as_index=False)["\u4ea7\u91cf(kg)"].sum()
    if rate_df.empty or not (pd.notna(lw_value) and lw_value > 0):
        return {}
    rate_df["_rate"] = rate_df["\u4ea7\u91cf(kg)"] / lw_value
    return {
        row["\u9879\u76ee"]: row["_rate"]
        for _, row in rate_df.iterrows()
        if pd.notna(row["_rate"]) and row["_rate"] > 0
    }


def _calc_restore_maps_for_day(overview, minors, df_lw, day_value, restore_mapping):
    ov_day = overview[overview["\u65e5\u671f"] == day_value][["\u9879\u76ee", "\u4ea7\u91cf(kg)", "\u542b\u7a0e\u91d1\u989d", "\u542b\u7a0e\u5355\u4ef7"]].copy()
    if ov_day.empty:
        return {}, {}, float("nan"), ov_day
    ov_day["\u9879\u76ee"] = ov_day["\u9879\u76ee"].map(_unify_restore_part)

    lw_day = _restore_lw_on(df_lw, day_value)
    if not (pd.notna(lw_day) and float(lw_day) > 0):
        return {}, {}, lw_day, ov_day

    rate_dict = _build_restore_rate_dict(ov_day, lw_day)
    if not rate_dict or minors is None or minors.empty or not restore_mapping:
        return {}, {}, lw_day, ov_day

    try:
        day_norm = pd.to_datetime(minors["\u65e5\u671f"], errors="coerce").dt.normalize()
        minors_day = minors.loc[day_norm == day_value].copy()
    except Exception:
        minors_day = pd.DataFrame()
    if minors_day.empty:
        return {}, {}, lw_day, ov_day

    code_col = "\u5b50\u7c7b" if "\u5b50\u7c7b" in minors_day.columns else None
    if code_col is None:
        for cand in ["\u54c1\u9879", "\u7269\u6599\u53f7", "\u540d\u79f0", "\u7269\u6599"]:
            if cand in minors_day.columns:
                code_col = cand
                break
    if code_col is None or "\u90e8\u4f4d\u5927\u7c7b" not in minors_day.columns:
        return {}, {}, lw_day, ov_day

    minors_day[code_col] = minors_day[code_col].astype(str).str.strip()
    minors_day["\u90e8\u4f4d\u5927\u7c7b"] = minors_day["\u90e8\u4f4d\u5927\u7c7b"].map(_unify_restore_part)
    grouped = minors_day.groupby(["\u90e8\u4f4d\u5927\u7c7b", code_col], as_index=False)["\u4ea7\u91cf(kg)"].sum()

    code_qty = {}
    code_part = {}
    for _, row in grouped.iterrows():
        code_raw = row[code_col]
        restore_part = row["\u90e8\u4f4d\u5927\u7c7b"]
        qty = float(row["\u4ea7\u91cf(kg)"]) if pd.notna(row["\u4ea7\u91cf(kg)"]) else 0.0
        if qty == 0:
            continue
        normed_code = normalize_code([code_raw])
        canon_code = normed_code[0] if normed_code else str(code_raw).strip()
        targets_raw = restore_mapping.get(canon_code, [])
        if not targets_raw:
            continue
        targets = [_unify_restore_part(t) for t in targets_raw if t]
        valid_targets = [t for t in targets if rate_dict.get(t, 0) > 0]
        if not valid_targets:
            continue
        code_qty[canon_code] = code_qty.get(canon_code, 0.0) + qty
        code_part[canon_code] = restore_part

    bone_rate_total = 0.0
    for code_k, qty_k in code_qty.items():
        if code_part.get(code_k) == "\u9aa8\u67b6\u7c7b" and qty_k:
            bone_rate_total += qty_k / lw_day

    inc_map = {}
    removed_map = {}
    for _, row in grouped.iterrows():
        restore_part = row["\u90e8\u4f4d\u5927\u7c7b"]
        code_raw = row[code_col]
        qty = float(row["\u4ea7\u91cf(kg)"]) if pd.notna(row["\u4ea7\u91cf(kg)"]) else 0.0
        if qty == 0:
            continue
        normed_code = normalize_code([code_raw])
        canon_code = normed_code[0] if normed_code else str(code_raw).strip()
        targets_raw = restore_mapping.get(canon_code, [])
        if not targets_raw:
            continue
        targets = [_unify_restore_part(t) for t in targets_raw if t]
        valid_targets = [t for t in targets if rate_dict.get(t, 0) > 0]
        if not valid_targets:
            continue

        use_bone_rule = (restore_part == "\u9aa8\u67b6\u7c7b")
        sum_target_rates = sum(rate_dict[t] for t in valid_targets)
        if use_bone_rule:
            adj_rates = {}
            for t in valid_targets:
                adj_rate = rate_dict[t]
                if t == "\u9aa8\u67b6\u7c7b":
                    adj_rate = adj_rate - bone_rate_total
                if adj_rate > 0:
                    adj_rates[t] = adj_rate
            denom = sum(adj_rates.values())
            if denom <= 0:
                continue
            removed_map[restore_part] = removed_map.get(restore_part, 0.0) + qty
            for t, adj_rate in adj_rates.items():
                share = adj_rate / denom
                inc = qty * share
                inc_map[t] = inc_map.get(t, 0.0) + inc
            continue

        if sum_target_rates <= 0:
            continue
        removed_map[restore_part] = removed_map.get(restore_part, 0.0) + qty
        base_qty = qty / sum_target_rates
        for t in valid_targets:
            rate_val = rate_dict[t]
            inc = base_qty * rate_val
            inc_map[t] = inc_map.get(t, 0.0) + inc

    return inc_map, removed_map, lw_day, ov_day


def _combine_restored_qty_map(qty_map):
    out = {str(k).strip(): float(v) for k, v in (qty_map or {}).items() if pd.notna(v)}
    chest_parts = ["\u80f8\u7c7b-\u80f8", "\u80f8\u7c7b-\u80f8\u76ae"]
    if any(p in out for p in chest_parts):
        out["\u80f8\u7c7b"] = sum(out.get(p, 0.0) for p in chest_parts)
    combo_parts = ["\u9e21\u5934\u7c7b", "\u8116\u7c7b", "\u9aa8\u67b6\u7c7b"]
    if any(p in out for p in combo_parts):
        out["\u9e21\u5934+\u9e21\u8116+\u9aa8\u67b6"] = sum(out.get(p, 0.0) for p in combo_parts)
    bone_neck_parts = ["\u9aa8\u67b6\u7c7b", "\u8116\u7c7b"]
    if any(p in out for p in bone_neck_parts):
        out["\u9aa8\u67b6+\u9e21\u8116"] = sum(out.get(p, 0.0) for p in bone_neck_parts)
    return out


def _compute_restored_main_side_qty_amt_for_period(overview, minors, df_lw, restore_mapping, start, end):
    if overview is None or overview.empty:
        return np.nan, np.nan, np.nan, np.nan

    mask = (overview["日期"] >= start) & (overview["日期"] <= end)
    seg = overview.loc[mask, ["日期", "项目", "产量(kg)", "含税金额"]].copy()
    if seg.empty:
        return np.nan, np.nan, np.nan, np.nan

    seg["项目"] = seg["项目"].map(_unify_restore_part)
    seg["日期"] = pd.to_datetime(seg["日期"], errors="coerce").dt.normalize()
    seg["产量(kg)"] = pd.to_numeric(seg["产量(kg)"], errors="coerce")
    seg["含税金额"] = pd.to_numeric(seg["含税金额"], errors="coerce")
    seg = seg.dropna(subset=["日期"])
    if seg.empty:
        return np.nan, np.nan, np.nan, np.nan

    dynamic_base = _dynamic_base_projects(seg["项目"])
    base_scope = seg[seg["项目"].isin(dynamic_base)].copy()
    if base_scope.empty:
        return np.nan, np.nan, np.nan, np.nan

    period_proj = base_scope.groupby("项目", as_index=False)[["产量(kg)", "含税金额"]].sum(min_count=1)
    qty_map = {}
    amt_map = {}
    unit_map = {}
    for _, row in period_proj.iterrows():
        proj = str(row["项目"]).strip()
        qty = float(row["产量(kg)"]) if pd.notna(row["产量(kg)"]) else 0.0
        amt = float(row["含税金额"]) if pd.notna(row["含税金额"]) else 0.0
        qty_map[proj] = qty
        amt_map[proj] = amt
        unit_map[proj] = (amt / qty) if qty != 0 else np.nan

    delta_acc = {}
    for day_value in sorted(base_scope["日期"].dropna().unique()):
        inc_map, removed_map, _, ov_day = _calc_restore_maps_for_day(overview, minors, df_lw, day_value, restore_mapping)
        if ov_day is None or ov_day.empty:
            continue
        for proj in (set(inc_map) | set(removed_map)):
            delta_acc[proj] = delta_acc.get(proj, 0.0) + float(inc_map.get(proj, 0.0)) - float(removed_map.get(proj, 0.0))

    restored_qty_map = {}
    for proj in (set(qty_map) | set(delta_acc)):
        restored_qty_map[proj] = float(qty_map.get(proj, 0.0)) + float(delta_acc.get(proj, 0.0))
    restored_qty_map = _combine_restored_qty_map(restored_qty_map)

    main_projects = _dynamic_main_projects(base_scope["项目"]) & dynamic_base
    total_qty = sum(float(qty_map.get(proj, 0.0)) for proj in dynamic_base)
    restored_total_qty = sum(float(restored_qty_map.get(proj, qty_map.get(proj, 0.0))) for proj in dynamic_base)
    total_qty_for_value = restored_total_qty if restored_total_qty > 0 else total_qty

    main_qty_total = sum(float(restored_qty_map.get(proj, qty_map.get(proj, 0.0))) for proj in main_projects)
    side_qty_total = (float(total_qty_for_value) - float(main_qty_total)) if pd.notna(total_qty_for_value) and pd.notna(main_qty_total) else np.nan

    main_amt_total = 0.0
    have_main_amt = False
    for proj in sorted(main_projects):
        unit = unit_map.get(proj, np.nan)
        qty_val = float(restored_qty_map.get(proj, qty_map.get(proj, 0.0)))
        if pd.notna(unit):
            have_main_amt = True
            main_amt_total += float(unit) * qty_val
    if not have_main_amt:
        main_amt_total = np.nan

    total_amt = sum(float(amt_map.get(proj, 0.0)) for proj in dynamic_base)
    side_amt_total = (float(total_amt) - float(main_amt_total)) if (pd.notna(total_amt) and pd.notna(main_amt_total)) else np.nan

    return (
        float(main_qty_total) if pd.notna(main_qty_total) else np.nan,
        float(main_amt_total) if pd.notna(main_amt_total) else np.nan,
        float(side_qty_total) if pd.notna(side_qty_total) else np.nan,
        float(side_amt_total) if pd.notna(side_amt_total) else np.nan,
    )


def _build_main_side_rows(
    seg,
    total_qty,
    total_amt,
    total_sale,
    overview_all=None,
    minors=None,
    df_lw=None,
    restore_mapping=None,
    start=None,
    end=None,
):
    columns = ["项目", "产量(kg)", "销量(kg)", "产成率%", "含税金额", "含税单价"]
    if seg is None or seg.empty:
        return pd.DataFrame(columns=columns)

    base_projects = _dynamic_base_projects(seg["项目"])
    main_projects = _dynamic_main_projects(seg["项目"])
    base_scope = seg[seg["项目"].isin(base_projects)].copy()
    if base_scope.empty:
        return pd.DataFrame(columns=columns)

    main_scope = base_scope[base_scope["项目"].isin(main_projects)].copy()
    side_scope = base_scope[~base_scope["项目"].isin(main_projects)].copy()

    def _sum_col(frame, col):
        if col not in frame.columns:
            return np.nan
        val = pd.to_numeric(frame[col], errors="coerce").sum(min_count=1)
        return float(val) if pd.notna(val) else np.nan

    main_qty = _sum_col(main_scope, "产量(kg)")
    main_amt = _sum_col(main_scope, "含税金额")
    main_sale = _sum_col(main_scope, "销量(kg)")
    side_qty = _sum_col(side_scope, "产量(kg)")
    side_amt = _sum_col(side_scope, "含税金额")
    side_sale_direct = _sum_col(side_scope, "销量(kg)")

    restored_main_qty = restored_main_amt = restored_side_qty = restored_side_amt = np.nan
    if (
        overview_all is not None
        and not getattr(overview_all, "empty", True)
        and start is not None
        and end is not None
    ):
        restored_main_qty, restored_main_amt, restored_side_qty, restored_side_amt = (
            _compute_restored_main_side_qty_amt_for_period(
                overview_all, minors, df_lw, restore_mapping, start, end
            )
        )

    if pd.notna(restored_main_qty):
        main_qty = float(restored_main_qty)
    if pd.notna(restored_main_amt):
        main_amt = float(restored_main_amt)
    if pd.notna(restored_side_qty):
        side_qty = float(restored_side_qty)
    if pd.notna(restored_side_amt):
        side_amt = float(restored_side_amt)

    if pd.notna(total_qty) and pd.notna(main_qty):
        side_qty = float(total_qty) - float(main_qty)
    if pd.notna(total_amt) and pd.notna(main_amt):
        side_amt = float(total_amt) - float(main_amt)

    if pd.notna(total_sale) and pd.notna(main_sale):
        side_sale = float(total_sale) - float(main_sale)
        if side_sale < -1e-9:
            side_sale = side_sale_direct
    else:
        side_sale = side_sale_direct

    total_qty_for_value = total_qty
    if not (pd.notna(total_qty_for_value) and total_qty_for_value > 0):
        parts = [float(v) for v in (main_qty, side_qty) if pd.notna(v)]
        total_qty_for_value = sum(parts) if parts else np.nan

    total_unit = (
        float(total_amt) / float(total_qty)
        if (pd.notna(total_amt) and pd.notna(total_qty) and float(total_qty) != 0.0)
        else np.nan
    )
    main_unit = _component_value_unit(total_qty_for_value, main_amt)
    side_unit = _side_component_value_unit(total_unit, main_unit, total_qty_for_value, side_amt)

    rows = pd.DataFrame(
        [
            {
                "项目": "主产品",
                "产量(kg)": main_qty,
                "销量(kg)": main_sale,
                "产成率%": np.nan,
                "含税金额": main_amt,
                "含税单价": main_unit,
            },
            {
                "项目": "副产品",
                "产量(kg)": side_qty,
                "销量(kg)": side_sale,
                "产成率%": np.nan,
                "含税金额": side_amt,
                "含税单价": side_unit,
            },
        ]
    )
    if "日期" in seg.columns:
        row_date = pd.to_datetime(start if start is not None else seg["日期"].iloc[0], errors="coerce")
        rows.insert(0, "日期", row_date)
    return rows


def apply_part_allocation(m_code_daily, alloc_long, code2major):
    """
    有‘日期’→只在该日生效；‘日期’空→默认通用（整月）。
    层级：按日 > 默认(通用) > 一对一映射。
    修复点：通用规则合并时去掉分摊表的“日期”列，避免合并后出现 日期_x/日期_y 导致 KeyError。
    """
    import pandas as pd, numpy as np
    if m_code_daily is None or m_code_daily.empty:
        return pd.DataFrame(columns=["日期","项目","物料号","产量(kg)","含税金额"])

    base = m_code_daily.copy()
    base["含税金额"] = base["产量(kg)"] * base["综合单价_filled"]

    if alloc_long is None or alloc_long.empty:
        base["项目"] = base["物料号"].map(code2major)
        return base[["日期","项目","物料号","产量(kg)","含税金额"]]

    day_alloc = alloc_long[alloc_long["日期"].notna()].copy()
    mon_alloc = alloc_long[alloc_long["日期"].isna()].copy()

    parts = []

    # 1) 按日分摊（精确匹配 日期+物料号）
    if not day_alloc.empty:
        a = base.merge(day_alloc, on=["日期","物料号"], how="inner")
        if not a.empty:
            a["产量(kg)"] = a["产量(kg)"] * a["权重"]
            a["含税金额"] = a["含税金额"] * a["权重"]
            parts.append(a[["日期","项目","物料号","产量(kg)","含税金额"]])

    # 2) 默认通用（对未覆盖的行，按 物料号 套用）
    if not mon_alloc.empty:
        # 关键：去掉 mon_alloc 的“日期”列，避免 merge 后出现 日期_x/日期_y
        mon_alloc_nodate = mon_alloc.drop(columns=["日期"], errors="ignore").copy()

        if parts:
            used = pd.concat(parts, ignore_index=True)[["日期","物料号"]].drop_duplicates()
            rem = base.merge(used, on=["日期","物料号"], how="left", indicator=True)
            rem = rem[rem["_merge"]=="left_only"].drop(columns="_merge")
        else:
            rem = base

        b = rem.merge(mon_alloc_nodate, on=["物料号"], how="inner")

        if "日期_x" in b.columns and "日期" not in b.columns:
            b = b.rename(columns={"日期_x":"日期"})
        if "日期_y" in b.columns:
            b = b.drop(columns=["日期_y"])

        if not b.empty:
            b["产量(kg)"] = b["产量(kg)"] * b["权重"]
            b["含税金额"] = b["含税金额"] * b["权重"]
            parts.append(b[["日期","项目","物料号","产量(kg)","含税金额"]])

    # 3) 仍未命中 → 回落到一对一映射
    if parts:
        covered = pd.concat(parts, ignore_index=True)[["日期","物料号"]].drop_duplicates()
        rest = base.merge(covered, on=["日期","物料号"], how="left", indicator=True)
        rest = rest[rest["_merge"]=="left_only"].drop(columns="_merge")
    else:
        rest = base

    if not rest.empty:
        rest["项目"] = rest["物料号"].map(code2major)
        parts.append(rest[["日期","项目","物料号","产量(kg)","含税金额"]])

    return pd.concat(parts, ignore_index=True) if parts else pd.DataFrame(
        columns=["日期","项目","物料号","产量(kg)","含税金额"]
    )


def build_overview_table_for_day(
    overview,
    sel,
    df_lw,
    total_qty_df=None,
    pr_raw=None,
    percent_as_text=True,
    round_values=True,
    minors=None,
    restore_mapping=None,
):
    """构造指定日期的“总览”展示表（含胸类合并、总计、产成率显示）。"""
    import pandas as pd, numpy as np
    if overview is None or overview.empty or sel is None:
        return pd.DataFrame()
    try:
        sel_dt = pd.to_datetime(sel, errors="coerce").normalize()
    except Exception:
        return pd.DataFrame()
    if pd.isna(sel_dt):
        return pd.DataFrame()

    ov_all = overview.copy()
    ov_all["_日期_norm"] = pd.to_datetime(ov_all["日期"], errors="coerce").dt.normalize()
    ov = ov_all[ov_all["_日期_norm"] == sel_dt].drop(columns="_日期_norm", errors="ignore")
    if ov.empty:
        return pd.DataFrame()

    must = pd.DataFrame({"日期": sel_dt, "项目": SPECIFIED})
    ov = must.merge(ov, on=["日期","项目"], how="left")
    for c in ["产量(kg)","销量(kg)","产成率%","含税金额","含税单价"]:
        if c not in ov.columns:
            ov[c] = np.nan
    ov["产量(kg)"] = ov["产量(kg)"].fillna(0.0)
    ov["销量(kg)"] = ov["销量(kg)"].fillna(0.0)
    ov["含税金额"] = ov["含税金额"].fillna(0.0)
    ov["含税单价"] = ov["含税单价"].fillna(0.0)

    others = overview[(pd.to_datetime(overview["日期"], errors="coerce").dt.normalize() == sel_dt) & (~overview["项目"].isin(SPECIFIED))].copy()
    try:
        others["项目"] = pd.Categorical(others["项目"], categories=[x for x in ORDER if x not in SPECIFIED], ordered=True)
        others = others.sort_values("项目")
    except Exception:
        pass
    if "销量(kg)" in others.columns:
        others["销量(kg)"] = others["销量(kg)"].fillna(0.0)

    base_today = pd.concat([ov, others], ignore_index=True)
    dynamic_base = set(BASE_FOR_TOTAL)
    present = set(base_today["项目"].dropna().astype(str).unique())
    if "胸类" in present:
        dynamic_base.discard("胸类-胸")
        dynamic_base.discard("胸类-胸皮")
        dynamic_base.add("胸类")
    else:
        dynamic_base.discard("胸类")
    base_today_no_combo = base_today[base_today["项目"].isin(dynamic_base)]
    tot_qty = base_today_no_combo["产量(kg)"].sum(min_count=1)
    # 总计-销量：按全量销量口径（销售净重 + 调理品数量取绝对值后汇总）
    # 总计-销量：全量销量口径（销售净重 + 调理品数量取绝对值），不依赖部位映射
    if total_qty_df is not None and not getattr(total_qty_df, "empty", True):
        tq = total_qty_df.copy()
        tq["日期"] = pd.to_datetime(tq["日期"], errors="coerce").dt.normalize()
        tot_sale = pd.to_numeric(tq.loc[tq["日期"] == sel_dt, "总销量(kg)"], errors="coerce").sum(min_count=1)
    else:
        tot_sale = base_today["销量(kg)"].sum(min_count=1) if "销量(kg)" in base_today.columns else np.nan
    tot_amt = base_today_no_combo["含税金额"].sum(min_count=1)
    tot_unit = (tot_amt / tot_qty) if (pd.notna(tot_qty) and tot_qty > 0) else 0.0

    main_side_rows = _build_main_side_rows(
        base_today,
        tot_qty,
        tot_amt,
        tot_sale,
        overview_all=overview,
        minors=minors,
        df_lw=df_lw,
        restore_mapping=restore_mapping,
        start=sel_dt,
        end=sel_dt,
    )

    total_row = pd.DataFrame({
        "日期": [sel_dt],
        "项目": ["总计"],
        "产量(kg)": [float(tot_qty) if pd.notna(tot_qty) else 0.0],
        "销量(kg)": [float(tot_sale) if pd.notna(tot_sale) else 0.0],
        "含税金额": [float(tot_amt) if pd.notna(tot_amt) else 0.0],
        "含税单价": [float(tot_unit)],
        "产成率%": [np.nan],
    })
    final_ov = pd.concat([base_today, main_side_rows, total_row], ignore_index=True)
    disp = _attach_rate_display(final_ov, df_lw).copy()
    for _c in ["产量(kg)","销量(kg)","产成率%","含税金额","含税单价"]:
        if _c in disp.columns:
            disp[_c] = pd.to_numeric(disp[_c], errors="coerce")
            if round_values:
                disp[_c] = disp[_c].round(2)
    if "日期" in disp.columns:
        disp["日期"] = pd.to_datetime(disp["日期"], errors="coerce").dt.strftime("%Y-%m-%d")
    _order_cols = ["日期", "项目", "产量(kg)", "销量(kg)", "产成率%", "含税金额", "含税单价"]
    disp = disp[[c for c in _order_cols if c in disp.columns]]
    if percent_as_text and "产成率%" in disp.columns:
        disp["产成率%"] = disp["产成率%"].apply(lambda v: "" if pd.isna(v) else f"{float(v):.2f}%")
    return disp

# ===== UI =====
st.sidebar.header("数据源（请上传）")
prod_upload = st.sidebar.file_uploader("产量", type=["xlsx","xls"])
sale_upload = st.sidebar.file_uploader("销量", type=["xlsx","xls"], accept_multiple_files=True)
trans_upload = st.sidebar.file_uploader("转调理品原料（可选）", type=["xlsx","xls"])
transfer_upload = st.sidebar.file_uploader("转调拨/调拨宫产量", type=["xlsx","xls"])
lw_upload    = st.sidebar.file_uploader("净重", type=["xlsx","xls","csv"])
shed_arrival_upload = st.sidebar.file_uploader("棚前到场（可选）", type=["xlsx","xls","csv"])
bom_upload   = st.sidebar.file_uploader("物料清单", type=["xlsx","xls","csv"])
manual_price = st.sidebar.file_uploader("补价表（可选上传）", type=["xlsx","xls","csv"])
alloc_file  = st.sidebar.file_uploader("部位分摊表（可选上传）", type=["xlsx","xls","csv"])

required_main = [("产量", prod_upload), ("销量", sale_upload), ("净重", lw_upload), ("物料清单", bom_upload)]
missing_main = [name for name, up in required_main if (up is None or (isinstance(up, list) and len(up) == 0))]
if missing_main:
    st.error(f"请上传：{', '.join(missing_main)} 文件。")
    st.stop()

xls, load_errors = build_main_sheet_bundle(prod_upload, sale_upload, trans_upload, transfer_upload)
manual_month_df = read_manual_month_price(manual_price)
df_lw      = read_liveweight(lw_upload)
shed_arrival_df = read_shed_arrival(shed_arrival_upload)
df_lw = merge_liveweight_with_arrival(df_lw, shed_arrival_df)
if shed_arrival_upload is not None and (shed_arrival_df is None or shed_arrival_df.empty):
    st.warning("已上传“棚前到场”文件，但未识别到“日期”或“棚前-到场”列，数据未生效。")
code2major, code2desc = read_bom_mapping(bom_upload)

alloc_long = read_part_allocation(alloc_file)

if load_errors:
    for msg in load_errors:
        st.warning(msg)
if xls is None:
    st.error("请上传完整的主数据文件（产量/销量）。")
    st.stop()
if df_lw is None or df_lw.empty:
    st.error("未上传或未读取到净重文件，无法继续。")
    st.stop()
if not code2major:
    st.error("未上传或未读取到物料清单，无法继续。")
    st.stop()

overview, minors = build_overview(xls, code2major, df_lw, manual_month_df, alloc_long)

# 导出所需表的占位
export_core_summary = pd.DataFrame()
export_daily_overview = pd.DataFrame()
export_month_cum = pd.DataFrame()
export_restore_day = pd.DataFrame()
export_restore_month = pd.DataFrame()
export_minors = pd.DataFrame()
export_month_code_detail = pd.DataFrame()

st.subheader("总览")
if overview is None or overview.empty:
    st.error("主数据中未读取到任何有效的产量记录，请检查“产量”工作表及列名（日期/物料/数量）。")
    st.stop()
else:
    # 选择日期优先使用销量口径日期；并按月展开为完整日历日（支持选择当月末）
    days = []
    try:
        _sale_days_df = build_daily_total_qty(xls)
        if _sale_days_df is not None and not getattr(_sale_days_df, "empty", True):
            _anchor_days = pd.to_datetime(_sale_days_df["日期"], errors="coerce").dropna().dt.normalize()
        else:
            _anchor_days = pd.to_datetime(overview["日期"], errors="coerce").dropna().dt.normalize()
    except Exception:
        _anchor_days = pd.to_datetime(overview["日期"], errors="coerce").dropna().dt.normalize()

    if len(_anchor_days) > 0:
        _months = sorted(_anchor_days.dt.to_period("M").unique())
        _all_days = []
        for _m in _months:
            _m_start = _m.to_timestamp(how="start").normalize()
            _m_end = _m.to_timestamp(how="end").normalize()
            _all_days.extend(list(pd.date_range(_m_start, _m_end, freq="D")))
        days = sorted(pd.to_datetime(pd.Series(_all_days), errors="coerce").dropna().dt.normalize().unique())
    else:
        days = sorted(pd.to_datetime(overview["日期"].dropna().unique()))
    sel  = st.selectbox("选择日期", days, index=len(days)-1 if days else 0, format_func=lambda d: pd.to_datetime(d).strftime("%Y-%m-%d"))

    # === 顶部关键指标汇总（当日 / 本月累计 / 上月全月） ===
    try:
        sel_dt = pd.to_datetime(sel).normalize()
        month_start = sel_dt.replace(day=1)
        prev_month_end = month_start - pd.Timedelta(days=1)
        prev_month_start = prev_month_end.replace(day=1)

        # 标准化净重表，便于区间聚合
        lw_dcol = next((c for c in ["日期","交鸡日期","记帐日期","记账日期","凭证日期","过账日期"] if c in df_lw.columns), None)
        lw_wcol = next((c for c in ["毛鸡净重(kg)","毛鸡净重","净重","净重(kg)"] if c in df_lw.columns), None)
        lw_ccol = next((c for c in ["交鸡量","交鸡只数","交鸡数量","鸡只数","只数","数量(只)","数量（只）","数量"] if c in df_lw.columns), None)
        lw_norm = None
        if lw_dcol and lw_wcol:
            lw_norm = df_lw[[lw_dcol, lw_wcol] + ([lw_ccol] if lw_ccol else [])].copy()
            lw_norm.columns = ["日期", "毛鸡净重(kg)"] + (["交鸡量"] if lw_ccol else [])
            lw_norm["日期"] = pd.to_datetime(lw_norm["日期"], errors="coerce").dt.normalize()
            lw_norm["毛鸡净重(kg)"] = pd.to_numeric(lw_norm["毛鸡净重(kg)"], errors="coerce")
            if lw_ccol:
                lw_norm["交鸡量"] = pd.to_numeric(lw_norm["交鸡量"], errors="coerce")
            lw_norm = lw_norm.dropna(subset=["日期","毛鸡净重(kg)"])

        def _lw_stats(start, end):
            if lw_norm is None or lw_norm.empty:
                return np.nan, np.nan, np.nan
            mask = (lw_norm["日期"] >= start) & (lw_norm["日期"] <= end)
            if not mask.any():
                return np.nan, np.nan, np.nan
            w_sum = lw_norm.loc[mask, "毛鸡净重(kg)"].sum(min_count=1)
            c_sum = lw_norm.loc[mask, "交鸡量"].sum(min_count=1) if "交鸡量" in lw_norm.columns else np.nan
            avg_w = (w_sum / c_sum) if (pd.notna(w_sum) and pd.notna(c_sum) and c_sum!=0) else np.nan
            return w_sum, c_sum, avg_w

        # 供核心指标概览使用的全量销量（按日汇总后取绝对值）
        total_qty_df = None
        try:
            total_qty_df = build_daily_total_qty(xls)
        except Exception:
            total_qty_df = None

        restore_mapping_state = read_restore_mapping_upload(st.session_state.get("restore_mapping_file"))

        def _summary_row(label, start, end):
            empty_row = {
                "\u65e5\u671f": label,
                "\u4ea7\u6210\u7387(%)": np.nan,
                "\u4ea7\u503c(\u5143/kg)": np.nan,
                "\u4e3b\u4ea7\u54c1\u4ea7\u6210\u7387(%)": np.nan,
                "\u4e3b\u4ea7\u54c1\u4ea7\u503c(\u5143/kg)": np.nan,
                "\u526f\u4ea7\u54c1\u4ea7\u6210\u7387(%)": np.nan,
                "\u526f\u4ea7\u54c1\u4ea7\u503c(\u5143/kg)": np.nan,
                "\u751f\u8089\u4ea7\u91cf(\u5428)": np.nan,
                "\u9500\u91cf(\u5428)": np.nan,
                "\u4ea7\u9500\u7387(%)": np.nan,
                "\u5bb0\u9e21\u91cf(\u5343\u53ea)": np.nan,
                "\u5747\u91cd(kg/\u53ea)": np.nan,
            }
            mask = (overview["\u65e5\u671f"] >= start) & (overview["\u65e5\u671f"] <= end)
            seg = overview.loc[mask].copy()
            if seg.empty:
                return empty_row

            lw_sum, c_sum, avg_w = _lw_stats(start, end)
            dynamic_base = _dynamic_base_projects(seg["\u9879\u76ee"])
            tot_qty, _, rate, unit = _segment_metrics(seg, dynamic_base, lw_sum)

            main_qty, main_amt, side_qty, side_amt = _compute_restored_main_side_qty_amt_for_period(
                overview, minors, df_lw, restore_mapping_state, start, end
            )
            main_rate = (main_qty / lw_sum * 100.0) if (pd.notna(lw_sum) and lw_sum > 0 and pd.notna(main_qty)) else np.nan
            side_rate = (side_qty / lw_sum * 100.0) if (pd.notna(lw_sum) and lw_sum > 0 and pd.notna(side_qty)) else np.nan
            restored_total_qty = tot_qty
            if not (pd.notna(restored_total_qty) and restored_total_qty != 0):
                restored_total_qty = (
                    (main_qty if pd.notna(main_qty) else 0.0) +
                    (side_qty if pd.notna(side_qty) else 0.0)
                )
            main_unit = _component_value_unit(restored_total_qty, main_amt)
            side_unit = _side_component_value_unit(unit, main_unit, restored_total_qty, side_amt)

            if total_qty_df is not None and not total_qty_df.empty:
                _mask = (total_qty_df["\u65e5\u671f"] >= start) & (total_qty_df["\u65e5\u671f"] <= end)
                tot_sale = pd.to_numeric(total_qty_df.loc[_mask, "\u603b\u9500\u91cf(kg)"], errors="coerce").sum(min_count=1)
            else:
                tot_sale = np.nan
            sale_rate = (tot_sale / tot_qty * 100.0) if (pd.notna(tot_qty) and tot_qty != 0 and pd.notna(tot_sale)) else np.nan

            return {
                "\u65e5\u671f": label,
                "\u4ea7\u6210\u7387(%)": rate,
                "\u4ea7\u503c(\u5143/kg)": unit,
                "\u4e3b\u4ea7\u54c1\u4ea7\u6210\u7387(%)": main_rate,
                "\u4e3b\u4ea7\u54c1\u4ea7\u503c(\u5143/kg)": main_unit,
                "\u526f\u4ea7\u54c1\u4ea7\u6210\u7387(%)": side_rate,
                "\u526f\u4ea7\u54c1\u4ea7\u503c(\u5143/kg)": side_unit,
                "\u751f\u8089\u4ea7\u91cf(\u5428)": (tot_qty / 1000.0) if pd.notna(tot_qty) else np.nan,
                "\u9500\u91cf(\u5428)": (tot_sale / 1000.0) if pd.notna(tot_sale) else np.nan,
                "\u4ea7\u9500\u7387(%)": sale_rate,
                "\u5bb0\u9e21\u91cf(\u5343\u53ea)": (c_sum / 1000.0) if pd.notna(c_sum) else np.nan,
                "\u5747\u91cd(kg/\u53ea)": avg_w,
            }

        summary_rows = [
            _summary_row(f"{sel_dt.month}月{sel_dt.day}日", sel_dt, sel_dt),
            _summary_row(f"{sel_dt.month}月累计", month_start, sel_dt),
            _summary_row(f"{prev_month_start.month}月全月累计", prev_month_start, prev_month_end)
        ]
        summary_df = pd.DataFrame(summary_rows)
        summary_num_cols = [c for c in summary_df.columns if c != "日期"]
        for col in summary_num_cols:
            if col in summary_df.columns:
                summary_df[col] = pd.to_numeric(summary_df[col], errors="coerce")

        export_core_summary = summary_df.copy()
        summary_disp = summary_df.copy()
        for col in summary_num_cols:
            if col in summary_disp.columns:
                summary_disp[col] = pd.to_numeric(summary_disp[col], errors="coerce").round(2)
        for col in [c for c in summary_disp.columns if "%" in str(c)]:
            if col in summary_disp.columns:
                summary_disp[col] = summary_disp[col].apply(lambda v: "" if pd.isna(v) else f"{float(v):.2f}%")
        st.markdown("### 核心指标概览")
        st.dataframe(summary_disp, use_container_width=True)
    except Exception as _e:
        st.warning(f"关键指标概览生成失败：{_e}")

    # 供当日总览使用的全量销量口径（按日汇总后取绝对值）
    total_qty_df = None
    try:
        total_qty_df = build_daily_total_qty(xls)
    except Exception:
        total_qty_df = None
    disp = build_overview_table_for_day(
        overview,
        sel,
        df_lw,
        total_qty_df=total_qty_df,
        percent_as_text=True,
        round_values=True,
        minors=minors,
        restore_mapping=restore_mapping_state,
    )
    export_daily_src = build_overview_table_for_day(
        overview,
        sel,
        df_lw,
        total_qty_df=total_qty_df,
        percent_as_text=False,
        round_values=False,
        minors=minors,
        restore_mapping=restore_mapping_state,
    )
    export_daily_overview = add_sale_rate(export_daily_src) if export_daily_src is not None else pd.DataFrame()
    if export_daily_overview is not None and not export_daily_overview.empty:
        export_daily_overview = export_daily_overview.drop(columns=["日期"], errors="ignore")
        if "产成率%" in export_daily_overview.columns:
            export_daily_overview["产成率%"] = pd.to_numeric(export_daily_overview["产成率%"], errors="coerce")
        if "产销率" in export_daily_overview.columns:
            export_daily_overview["产销率"] = pd.to_numeric(export_daily_overview["产销率"], errors="coerce") * 100.0
    st.dataframe(disp, use_container_width=True)

    # === 毛鸡净重 / 毛鸡均重（当日 & 当月） ===
    with st.expander("毛鸡净重 / 毛鸡均重", expanded=True):
        dcol = next((c for c in ["日期","交鸡日期","记帐日期","记账日期","凭证日期","过账日期"] if c in df_lw.columns), None)
        wcol = next((c for c in ["毛鸡净重(kg)","毛鸡净重","净重","净重(kg)"] if c in df_lw.columns), None)
        ccol = next((c for c in ["交鸡量","交鸡只数","交鸡数量","鸡只数","只数","数量(只)","数量（只）","数量"] if c in df_lw.columns), None)
        if not (dcol and wcol):
            st.warning("净重表缺少“日期”或“净重”列，无法展示毛鸡净重。")
        else:
            lw_calc = df_lw[[dcol, wcol] + ([ccol] if ccol else [])].copy()
            lw_calc[dcol] = pd.to_datetime(lw_calc[dcol], errors="coerce").dt.normalize()
            lw_calc[wcol] = pd.to_numeric(lw_calc[wcol], errors="coerce")
            if ccol:
                lw_calc[ccol] = pd.to_numeric(lw_calc[ccol], errors="coerce")
            lw_calc = lw_calc[lw_calc[dcol].notna() & lw_calc[wcol].notna()]

            sel_day = pd.to_datetime(sel).normalize()
            month_start = sel_day.replace(day=1)

            day_mask = lw_calc[dcol] == sel_day
            month_mask = (lw_calc[dcol] >= month_start) & (lw_calc[dcol] <= sel_day)

            day_weight = lw_calc.loc[day_mask, wcol].sum(min_count=1)
            month_weight = lw_calc.loc[month_mask, wcol].sum(min_count=1)

            day_count = lw_calc.loc[day_mask, ccol].sum(min_count=1) if ccol else np.nan
            month_count = lw_calc.loc[month_mask, ccol].sum(min_count=1) if ccol else np.nan

            day_avg = (day_weight / day_count) if (ccol and pd.notna(day_count) and day_count>0) else np.nan
            month_avg = (month_weight / month_count) if (ccol and pd.notna(month_count) and month_count>0) else np.nan

            display_rows = [
                {"口径": "当日", "毛鸡净重(kg)": day_weight, "交鸡量": day_count, "毛鸡均重(kg/只)": day_avg},
                {"口径": "当月", "毛鸡净重(kg)": month_weight, "交鸡量": month_count, "毛鸡均重(kg/只)": month_avg},
            ]
            disp_lw = pd.DataFrame(display_rows)
            for col in ["毛鸡净重(kg)", "交鸡量", "毛鸡均重(kg/只)"]:
                if col in disp_lw.columns:
                    disp_lw[col] = disp_lw[col].apply(lambda v: np.nan if pd.isna(v) else round(float(v), 2))

            show_cols = ["口径","毛鸡净重(kg)"]
            if ccol:
                show_cols += ["交鸡量","毛鸡均重(kg/只)"]
            else:
                st.info("净重表未提供“交鸡量”，仅显示净重，均重无法计算。")
                disp_lw = disp_lw.drop(columns=["交鸡量","毛鸡均重(kg/只)"], errors="ignore")
            st.dataframe(disp_lw[show_cols], use_container_width=True)

    # === 累计（本月至所选日） ===
    with st.expander("累计（本月至所选日）", expanded=True):
        sel_dt = pd.to_datetime(sel).normalize()
        month_start = sel_dt.replace(day=1)

        # 选取本月从1号到所选日（含）的区间
        rng = (overview["日期"] >= month_start) & (overview["日期"] <= sel_dt)

        # 当月销量（仅保留展示用）
        sales_cum = pd.DataFrame()
        if "销量(kg)" in overview.columns:
            try:
                sales_cum = (overview.loc[rng]
                                    .groupby("项目", as_index=False)["销量(kg)"]
                                    .sum(min_count=1))
            except Exception:
                sales_cum = pd.DataFrame()

        # —— 物料月均价：当月收入 / 当月数量（含转调1.09） ——
        price_code_mtd = pd.DataFrame(columns=["物料号","月均价"])
        try:
            pr_daily = build_daily_code_price_raw(xls)
            if pr_daily is not None and not getattr(pr_daily, "empty", True):
                pr_use = pr_daily.copy()
                pr_use["日期"] = pd.to_datetime(pr_use["日期"], errors="coerce").dt.normalize()
                pr_use = pr_use[(pr_use["日期"] >= month_start) & (pr_use["日期"] <= sel_dt)]
                pr_use["金额"] = pd.to_numeric(pr_use["金额"], errors="coerce")
                pr_use["数量"] = pd.to_numeric(pr_use["数量"], errors="coerce")
                pr_use = pr_use[pr_use["数量"].notna() & (pr_use["数量"] != 0)]
                if not pr_use.empty:
                    pr_month = pr_use.groupby("物料号", as_index=False)[["金额","数量"]].sum(min_count=1)
                    pr_month["月均价"] = np.where(pr_month["数量"] != 0, pr_month["金额"]/pr_month["数量"], np.nan)
                    price_code_mtd = pr_month[["物料号","月均价"]]
            # 补价表兜底月均价
            if manual_month_df is not None and not getattr(manual_month_df, "empty", True):
                mp = manual_month_df.copy()
                mp["物料号"] = mp["物料号"].astype(str).str.strip()
                mp = mp.rename(columns={"手工单价":"_手工单价"})
                price_code_mtd = price_code_mtd.merge(mp[["物料号","_手工单价"]], on="物料号", how="outer")
                price_code_mtd["月均价"] = price_code_mtd["月均价"].where(price_code_mtd["月均价"].notna(), price_code_mtd["_手工单价"])
                price_code_mtd = price_code_mtd.drop(columns=["_手工单价"], errors="ignore")
        except Exception:
            price_code_mtd = pd.DataFrame(columns=["物料号","月均价"])

        # —— 当月产量（已分摊后的子类） × 月均价 → 金额，再按部位汇总 ——
        cum_base = pd.DataFrame()
        try:
            if minors is not None and not minors.empty:
                minor_month = minors.copy()
                minor_month["日期"] = pd.to_datetime(minor_month["日期"], errors="coerce").dt.normalize()
                mask_minor = (minor_month["日期"] >= month_start) & (minor_month["日期"] <= sel_dt)
                minor_month = minor_month.loc[mask_minor].rename(columns={"子类": "物料号", "部位大类": "项目"})
                if not minor_month.empty:
                    minor_month["产量(kg)"] = pd.to_numeric(minor_month["产量(kg)"], errors="coerce").fillna(0.0)
                    detail = minor_month.groupby(["项目", "物料号"], as_index=False)[["产量(kg)"]].sum(min_count=1)
                    detail = detail.merge(price_code_mtd, on="物料号", how="left")
                    detail["含税金额"] = detail["月均价"] * detail["产量(kg)"]
                    proj_agg = detail.groupby("项目", as_index=False).agg({
                        "产量(kg)": "sum",
                        "含税金额": "sum"
                    })
                    # 组合项目（胸类、鸡头+鸡脖+骨架）
                    frames = [proj_agg]
                    chest = proj_agg[proj_agg["项目"].isin(["胸类-胸","胸类-胸皮"])]
                    if not chest.empty:
                        s = chest[["产量(kg)","含税金额"]].sum(min_count=1)
                        frames.append(pd.DataFrame([{
                            "项目": "胸类",
                            "产量(kg)": s["产量(kg)"],
                            "含税金额": s["含税金额"],
                        }]))
                    comb = proj_agg[proj_agg["项目"].isin(["鸡头类","脖类","骨架类"])]
                    if not comb.empty:
                        s = comb[["产量(kg)","含税金额"]].sum(min_count=1)
                        frames.append(pd.DataFrame([{
                            "项目": "鸡头+鸡脖+骨架",
                            "产量(kg)": s["产量(kg)"],
                            "含税金额": s["含税金额"],
                        }]))
                    cum_base = pd.concat(frames, ignore_index=True)
        except Exception:
            cum_base = pd.DataFrame()

        # 必显顺序 + 补零
        must_c = pd.DataFrame({"项目": SPECIFIED})
        cum = must_c.merge(cum_base, on="项目", how="left")
        for c in ["产量(kg)","含税金额"]:
            if c not in cum.columns:
                cum[c] = np.nan
        if not sales_cum.empty:
            cum = cum.merge(sales_cum, on="项目", how="left")
        if "销量(kg)" not in cum.columns:
            cum["销量(kg)"] = np.nan
        cum["产量(kg)"] = pd.to_numeric(cum["产量(kg)"], errors="coerce").fillna(0.0)
        cum["销量(kg)"] = pd.to_numeric(cum["销量(kg)"], errors="coerce").fillna(0.0)
        cum["含税金额"] = pd.to_numeric(cum["含税金额"], errors="coerce").fillna(0.0)
        cum["含税单价"] = np.where(cum["产量(kg)"] != 0, cum["含税金额"]/cum["产量(kg)"], np.nan)

        # 其它大类（非必显）
        others_c = pd.DataFrame()
        if cum_base is not None and not cum_base.empty:
            others_c = cum_base[~cum_base["项目"].isin(SPECIFIED)].copy()
            if not others_c.empty:
                others_c = others_c.merge(sales_cum, on="项目", how="left")
                others_c["销量(kg)"] = pd.to_numeric(others_c["销量(kg)"], errors="coerce").fillna(0.0)
                others_c["含税单价"] = np.where(others_c["产量(kg)"] != 0, others_c["含税金额"]/others_c["产量(kg)"], np.nan)
                try:
                    others_c["项目"] = pd.Categorical(others_c["项目"], categories=[x for x in ORDER if x not in SPECIFIED], ordered=True)
                    others_c = others_c.sort_values("项目")
                except Exception:
                    pass
        cum_base = pd.concat([cum, others_c], ignore_index=True) if 'others_c' in locals() else cum

        # 总计（仅基础大类）
        dynamic_base_cum = set(BASE_FOR_TOTAL)
        # —— 互斥：若区间含“胸类”，则排除“胸类-胸/胸类-胸皮”；否则排除“胸类” ——
        present_cum = set(cum_base['项目'].dropna().astype(str).unique())
        if '胸类' in present_cum:
            dynamic_base_cum.discard('胸类-胸')
            dynamic_base_cum.discard('胸类-胸皮')
            dynamic_base_cum.add('胸类')
        else:
            dynamic_base_cum.discard('胸类')
        base_mask = cum_base['项目'].isin(dynamic_base_cum)
        tot_qty = cum_base.loc[base_mask, '产量(kg)'].sum(min_count=1)
        # 总计-销量：全量销量口径（按日汇总后取绝对值），不依赖部位映射
        if total_qty_df is not None and not total_qty_df.empty:
            _mask = (total_qty_df["日期"] >= month_start) & (total_qty_df["日期"] <= sel_dt)
            tot_sale = pd.to_numeric(total_qty_df.loc[_mask, "总销量(kg)"], errors="coerce").sum(min_count=1)
        else:
            tot_sale = np.nan
        tot_amt = cum_base.loc[base_mask, '含税金额'].sum(min_count=1)
        tot_unit = (tot_amt/tot_qty) if (pd.notna(tot_qty) and tot_qty>0) else 0.0

        main_side_rows = _build_main_side_rows(
            cum_base,
            tot_qty,
            tot_amt,
            tot_sale,
            overview_all=overview,
            minors=minors,
            df_lw=df_lw,
            restore_mapping=restore_mapping_state,
            start=month_start,
            end=sel_dt,
        )
        total_row = pd.DataFrame({"项目":["总计"],                               "产量(kg)":[float(tot_qty) if pd.notna(tot_qty) else 0.0],                               "销量(kg)":[float(tot_sale) if pd.notna(tot_sale) else 0.0],                               "含税金额":[float(tot_amt) if pd.notna(tot_amt) else 0.0],                               "含税单价":[float(tot_unit)]})
        cum_final = pd.concat([cum_base, main_side_rows, total_row], ignore_index=True)

        # 可选：累计产成率（以本月毛鸡净重累计为分母）
        try:
            lw_sum = None
            if df_lw is not None and not df_lw.empty:
                dcol = next((c for c in ["日期","交鸡日期","记帐日期","记账日期","凭证日期","过账日期"] if c in df_lw.columns), None)
                vcol = next((c for c in ["毛鸡净重(kg)","毛鸡净重","净重","净重(kg)"] if c in df_lw.columns), None)
                if dcol and vcol:
                    _lw = df_lw[[dcol, vcol]].copy()
                    _lw.columns = ["日期","_lw"]
                    _lw["日期"] = pd.to_datetime(_lw["日期"], errors="coerce").dt.normalize()
                    lw_sum = _lw.loc[(_lw["日期"]>=month_start)&(_lw["日期"]<=sel_dt), "_lw"].sum()
            if lw_sum and lw_sum>0:
                cum_final["产成率%"] = (cum_final["产量(kg)"] / lw_sum) * 100.0
        except Exception:
            pass

        for _c in ['产量(kg)','销量(kg)','产成率%','含税金额','含税单价']:
            if _c in cum_final.columns:
                cum_final[_c] = pd.to_numeric(cum_final[_c], errors="coerce")

        order_cols_cum = ["项目","产量(kg)","销量(kg)","产成率%","含税金额","含税单价"]
        cum_final = cum_final[[c for c in order_cols_cum if c in cum_final.columns]]
        export_month_cum = add_sale_rate(cum_final)
        if "产销率" in export_month_cum.columns:
            export_month_cum["产销率"] = pd.to_numeric(export_month_cum["产销率"], errors="coerce") * 100.0

        cum_disp = cum_final.copy()
        for _c in ['产量(kg)','销量(kg)','产成率%','含税金额','含税单价']:
            if _c in cum_disp.columns:
                cum_disp[_c] = pd.to_numeric(cum_disp[_c], errors="coerce").round(2)
        if "产成率%" in cum_disp.columns:
            cum_disp["产成率%"] = cum_disp["产成率%"].apply(
                lambda v: "" if pd.isna(v) else f"{float(v):.2f}%"
            )
        st.dataframe(cum_disp, use_container_width=True)

        # Align core summary cumulative unit value with cumulative total row.
        try:
            month_label = f"{sel_dt.month}\u6708\u7d2f\u8ba1"
            if (
                export_core_summary is not None and not export_core_summary.empty
                and export_month_cum is not None and not export_month_cum.empty
                and {"\u65e5\u671f", "\u4ea7\u503c(\u5143/kg)"} <= set(export_core_summary.columns)
            ):
                mask = export_core_summary["\u65e5\u671f"] == month_label
                if mask.any():
                    total_row = export_month_cum[export_month_cum["\u9879\u76ee"] == "\u603b\u8ba1"]
                    if not total_row.empty and "\u542b\u7a0e\u5355\u4ef7" in total_row.columns:
                        unit_val = pd.to_numeric(total_row.iloc[0]["\u542b\u7a0e\u5355\u4ef7"], errors="coerce")
                        if pd.notna(unit_val):
                            export_core_summary.loc[mask, "\u4ea7\u503c(\u5143/kg)"] = float(unit_val)
        except Exception:
            pass

        if minors is not None and not minors.empty:
            try:
                minor_month = minors.copy()
                minor_month["日期"] = pd.to_datetime(minor_month["日期"], errors="coerce").dt.normalize()
                mask_minor = (minor_month["日期"] >= month_start) & (minor_month["日期"] <= sel_dt)
                minor_month = minor_month.loc[mask_minor].rename(columns={"子类": "物料号"})
                if not minor_month.empty:
                    minor_month["产量(kg)"] = pd.to_numeric(minor_month["产量(kg)"], errors="coerce").fillna(0.0)
                    # 复用当月月均价（绝对收入/绝对数量）口径
                    code_price = price_code_mtd if 'price_code_mtd' in locals() else pd.DataFrame(columns=["物料号","月均价"])
                    code_detail = (minor_month
                                   .groupby(["部位大类", "物料号"], as_index=False)[["产量(kg)"]]
                                   .sum(min_count=1))
                    code_detail = code_detail.merge(code_price, on="物料号", how="left")
                    code_detail["含税金额"] = code_detail["月均价"] * code_detail["产量(kg)"]
                    if code2desc:
                        code_detail["物料描述"] = code_detail["物料号"].map(code2desc)
                    try:
                        code_detail = code_detail.sort_values(["部位大类", "物料号"])
                    except Exception:
                        pass
                    cols_order = ["部位大类", "物料号", "物料描述", "产量(kg)", "含税金额", "月均价"]
                    code_detail = code_detail[[c for c in cols_order if c in code_detail.columns]]
                    for _c in ["产量(kg)", "含税金额", "月均价"]:
                        if _c in code_detail.columns:
                            code_detail[_c] = pd.to_numeric(code_detail[_c], errors="coerce")
                    display_code_detail = format_two_decimals(
                        format_thousands(code_detail.copy(), ["产量(kg)", "含税金额"]),
                        ["月均价"]
                    )
                    export_month_code_detail = code_detail.copy()
                    st.markdown("**物料月均价明细（本月至所选日）**")
                    st.dataframe(display_code_detail, use_container_width=True)
            except Exception:
                st.warning("物料月均价明细生成失败，请检查数据。")


if minors is not None and not minors.empty:
    st.subheader("子类")
    minors_export = minors.copy()
    if code2desc:
        minors_export["物料描述"] = minors_export["子类"].map(code2desc)
    export_minors = minors_export.copy()
    order_cols_minors = ["日期","部位大类","子类","物料描述","产量(kg)","含税金额","含税单价"]
    if "日期" in export_minors.columns:
        export_minors["日期"] = pd.to_datetime(export_minors["日期"], errors="coerce").dt.strftime("%m-%d")
    for _c in ["产量(kg)", "含税金额", "含税单价", "产成率%"]:
        if _c in export_minors.columns:
            export_minors[_c] = pd.to_numeric(export_minors[_c], errors="coerce")
    export_minors = export_minors[[c for c in order_cols_minors if c in export_minors.columns]]
    days2 = sorted(pd.to_datetime(minors["日期"].dropna().unique()))
    sel2  = st.selectbox("选择日期（下钻）", days2, index=len(days2)-1 if days2 else 0, format_func=lambda d: pd.to_datetime(d).strftime("%Y-%m-%d"))
    for mj in ORDER:
        if mj in ("总计",): continue
        sub = minors[(minors["日期"]==sel2) & (minors["部位大类"]==mj)]
        if sub.empty: continue
        with st.expander(f"{mj} 子类明细", expanded=False):
            sub_disp = sub.copy()
            if code2desc:
                sub_disp["物料描述"] = sub_disp["子类"].map(code2desc)
            for _c in ['产量(kg)','产成率%','含税金额','含税单价']:
                if _c in sub_disp.columns:
                    sub_disp[_c] = pd.to_numeric(sub_disp[_c], errors="coerce").round(2)
            if "日期" in sub_disp.columns:
                sub_disp["日期"] = pd.to_datetime(sub_disp["日期"], errors="coerce").dt.strftime("%Y-%m-%d")
            order_cols = ["日期","部位大类","子类","物料描述","产量(kg)","产成率%","含税金额","含税单价"]
            sub_disp = sub_disp[[c for c in order_cols if c in sub_disp.columns]]
            st.dataframe(sub_disp, use_container_width=True)

# === 概览报表下载 ===
try:
    th_cols = ["产量(kg)","销量(kg)","含税金额","原产量(kg)","调整量(kg)","调整后产量(kg)"]
    # 当日总览允许缺失：缺失时仍可下载，导出时自动跳过该分段
    have_overview = all([
        export_core_summary is not None and not export_core_summary.empty,
        export_month_cum is not None and not export_month_cum.empty,
    ])
    if have_overview:
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            sheet_name = "报表"
            start_row = 0

            sections = [
                ("核心指标概览", export_core_summary),
                ("当日总览", export_daily_overview),
                ("本月至今累计", export_month_cum),
            ]

            for title, df_sec in sections:
                if df_sec is None or df_sec.empty:
                    continue
                if title == "核心指标概览":
                    start_row += _write_core_summary_excel_section(writer, sheet_name, start_row, df_sec, title)
                    continue
                df_sec.to_excel(writer, sheet_name=sheet_name, index=False, startrow=start_row+1)
                ws = writer.sheets.get(sheet_name)
                if ws is None:
                    ws = writer.book[sheet_name]
                ws.cell(row=start_row+1, column=1, value=title)
                _apply_excel_formats_by_header(
                    ws=ws,
                    header_row=start_row + 2,
                    data_start_row=start_row + 3,
                    data_end_row=start_row + 2 + len(df_sec),
                    col_format_map=_build_excel_col_format_map(df_sec.columns),
                )
                start_row += len(df_sec) + 3  # 标题行 + 数据 + 间隔行

            if export_minors is not None and not export_minors.empty:
                export_minors.to_excel(writer, sheet_name="SKU价格明细", index=False)
                ws_m = writer.sheets.get("SKU价格明细")
                _apply_excel_formats_by_header(
                    ws=ws_m,
                    header_row=1,
                    data_start_row=2,
                    data_end_row=1 + len(export_minors),
                    col_format_map=_build_excel_col_format_map(export_minors.columns),
                )
            if export_month_code_detail is not None and not export_month_code_detail.empty:
                export_month_code_detail.to_excel(writer, sheet_name="物料月均价明细", index=False)
                ws_d = writer.sheets.get("物料月均价明细")
                _apply_excel_formats_by_header(
                    ws=ws_d,
                    header_row=1,
                    data_start_row=2,
                    data_end_row=1 + len(export_month_code_detail),
                    col_format_map=_build_excel_col_format_map(export_month_code_detail.columns),
                )

        st.download_button(
            "下载概览报表（XLSX）",
            data=buffer.getvalue(),
            file_name="概览报表.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        missing = []
        if export_core_summary is None or export_core_summary.empty:
            missing.append("核心指标概览")
        if export_month_cum is None or export_month_cum.empty:
            missing.append("本月至今累计")
        if missing:
            st.info("请先生成概览数据后再下载（缺少：" + "、".join(missing) + "）。")
except Exception as _e:
    st.error(f"概览报表下载生成失败：{_e}")

# === 新增：当天“无售价”的物料号清单（仅提示，不改计算；避免嵌套 expander） ===
st.divider()
st.markdown("#### 当月无售价的物料号")
_show_missing = st.checkbox("显示清单", value=True)

if _show_missing:
    if overview is None or overview.empty:
        st.warning("没有总览数据，无法生成“当月无售价”的清单。")
    else:
        # 全量（当月范围）
        qty_all = read_qty_per_code_per_day(xls)
        pr_all  = build_daily_code_price_raw(xls)

        # 选定参考月份：优先 sel / sel2，其次最新日期
        sel_month = resolve_sel_month(overview)
        if sel_month is None:
            st.warning("无法识别可用日期，无法生成“当月无售价”的清单。")
            st.stop()

        qd  = qty_all.loc[pd.to_datetime(qty_all["日期"]).dt.to_period("M") == sel_month].copy()
        pd0 = pr_all .loc[pd.to_datetime(pr_all ["日期"]).dt.to_period("M") == sel_month].copy()

        # 容错：缺列时补空列；统一数值类型
        if "综合单价" not in pd0.columns:
            pd0["综合单价"] = np.nan
        if "数量" not in pd0.columns:
            pd0["数量"] = 0

        pd0["综合单价"] = pd.to_numeric(pd0["综合单价"], errors="coerce")
        pd0["数量"]    = pd.to_numeric(pd0["数量"],    errors="coerce").fillna(0)

        # “有售价”口径（保持原口径不变）：数量≠0 且 单价存在且>0
        has_price_mask  = (pd0["数量"] != 0) & pd0["综合单价"].notna() & (pd0["综合单价"] > 0)
        # 在本月任一日有售价即视为“本月有售价”
        price_codes_month = set(pd0.loc[has_price_mask, "物料号"].dropna().astype(str).unique())

        # 在产量里出现、但整月没有任何有效售价的物料号
        qd["物料号"] = qd["物料号"].astype(str)
        _missing = (
            qd[["物料号"]].drop_duplicates()
              .loc[lambda d: ~d["物料号"].isin(price_codes_month)]
        )

        if _missing.empty:
            st.success("✅ 当月所有参与产量的物料号均至少有一次有效售价记录。")
        else:
            # 展示：部位映射 & 本月合计产量
            miss = (
                _missing
                .merge(qd.groupby("物料号", as_index=False)["产量(kg)"].sum(), on="物料号", how="left")
            )
            try:
                miss["部位"] = miss["物料号"].map(code2major).fillna("未映射")
            except Exception:
                miss["部位"] = "未映射"

            miss = (
                miss.assign(_ord=lambda d: (d["部位"] == "未映射").astype(int))
                    .sort_values(["_ord", "产量(kg)"], ascending=[False, False])
                    .drop(columns="_ord")
            )
            miss["产量(kg)"] = pd.to_numeric(miss["产量(kg)"], errors="coerce").round(2)

            if code2desc:
                miss["物料描述"] = miss["物料号"].map(code2desc)

            display_cols = ["物料号", "物料描述", "部位", "产量(kg)"]
            st.dataframe(miss[[c for c in display_cols if c in miss.columns]], use_container_width=True)
# === 新增窗口：自定义部位还原（同日） + 分布后总览 ===
try:
    st.subheader("部位还原")

    if overview is None or (hasattr(overview, "empty") and overview.empty):
        st.info("无数据，无法计算。")
    else:
        _alias_extra = {
            "腿":"腿类","胸":"胸类","里肌":"里肌类","翅":"翅类",
            "鸡骨架":"骨架类","骨架":"骨架类",
            "鸡头":"鸡头类","鸡肝":"鸡肝类","鸡心":"鸡心类",
            "脖":"脖类","鸡脖":"脖类",
            "整鸡":"整鸡类","其他内脏":"其他内脏","下料":"下料类","下料类":"下料类"
        }

        def _unify(name: str) -> str:
            x = str(name)
            try:
                if 'ALIAS' in globals() and x in ALIAS:
                    return ALIAS[x]
            except Exception:
                pass
            return _alias_extra.get(x, x)

        def _parse_mapping(text_value: str):
            mapping = {}
            for line in text_value.splitlines():
                row = line.strip()
                if not row or row.startswith(('#', '/')):
                    continue
                token = None
                if ':' in row:
                    token = ':'
                elif '->' in row:
                    token = '->'
                if token is None:
                    continue
                code, targets = row.split(token, 1)
                code = code.strip()
                targets = targets.strip()
                if not code or not targets:
                    continue
                pieces = [t.strip() for t in re.split(r"[，,;；/|]", targets) if t.strip()]
                if pieces:
                    mapping[code] = pieces
            return mapping

        def _read_restore_mapping(uploaded):
            """读取“物料号→目标部位”映射表，支持 xlsx/xls/csv，两列：物料号、目标部位。"""
            mapping = {}
            if uploaded is None:
                return mapping
            try:
                frames = []
                name = _get_name(uploaded).lower()
                if name.endswith(".csv"):
                    raw = _load_raw(uploaded)
                    if raw is None:
                        return mapping
                    frames = [pd.read_csv(io.BytesIO(raw))]
                else:
                    xls = _load_excel(uploaded)
                    if xls is None:
                        return mapping
                    for s in _get_sheet_names(xls):
                        df = _read_sheet_df(xls, s)
                        if df is not None:
                            frames.append(df)
                for df in frames:
                    cols = [str(c).strip() for c in df.columns]
                    code_col = next((c for c in cols if any(k in c for k in ["物料号","物料编码","物料编号","物料代码","编码","Material","品项","物料"])), None)
                    tgt_col  = next((c for c in cols if any(k in c for k in ["目标","还原","去向","分配到","部位"])), None)
                    if not (code_col and tgt_col):
                        continue
                    subset = df[[code_col, tgt_col]].dropna(subset=[code_col, tgt_col]).copy()
                    for _, row in subset.iterrows():
                        codes = normalize_code([row[code_col]])
                        if not codes:
                            continue
                        code = codes[0]
                        tgt_raw = str(row[tgt_col]).strip()
                        if not tgt_raw:
                            continue
                        pieces = [t.strip() for t in re.split(r"[，,;；/|、]+", tgt_raw) if t.strip()]
                        if not pieces:
                            continue
                        existing = mapping.get(code, [])
                        for p in pieces:
                            if p not in existing:
                                existing.append(p)
                        mapping[code] = existing
                return mapping
            except Exception:
                return {}


        _all_days = sorted(pd.to_datetime(overview["日期"].dropna().unique()))
        if not _all_days:
            st.info("无可用日期。")
            st.stop()
        _latest = max(_all_days)
        _latest_m = pd.Timestamp(_latest).to_period("M")
        _month_days = [d for d in _all_days if pd.Timestamp(d).to_period("M") == _latest_m]
        _first = [d for d in _month_days if pd.Timestamp(d).day == 1]
        _default_ref = _first[0] if _first else _month_days[0]

        ref_day = st.selectbox(
            "还原日（同日口径）",
            _all_days,
            index=_all_days.index(_default_ref),
            format_func=lambda d: pd.to_datetime(d).strftime("%Y-%m-%d")
        )
        ref_day = pd.to_datetime(ref_day).normalize()

        proj_all = [p for p in (ORDER if 'ORDER' in globals() else sorted(overview["项目"].dropna().unique().tolist())) if p != "总计"]

        restore_mapping_file = st.file_uploader(
            "物料还原映射表（可选上传）",
            type=["xlsx","xls","csv"],
            help="两列：物料号、目标部位；多个目标可用逗号/分号/斜杠分隔。",
            key="restore_mapping_file"
        )
        restore_mapping = _read_restore_mapping(restore_mapping_file)
        if restore_mapping_file is not None:
            msg = f"已读取 {len(restore_mapping)} 个物料号的还原去向。" if restore_mapping else "未在映射表中识别到有效的列/数据。"
            st.caption(msg)
        if not restore_mapping:
            st.warning("请上传包含“物料号→目标部位”的映射表后再继续。")
            st.stop()

        def _canon_code_val(v):
            normed = normalize_code([v])
            return normed[0] if normed else str(v).strip()

        def _lw_on(day_value):
            val = float("nan")
            try:
                if df_lw is not None and not df_lw.empty:
                    dcol = next((c for c in ["日期","交鸡日期","记帐日期","记账日期","凭证日期","过账日期"] if c in df_lw.columns), None)
                    vcol = next((c for c in ["毛鸡净重(kg)","毛鸡净重","净重","净重(kg)"] if c in df_lw.columns), None)
                    if dcol and vcol:
                        _lw = df_lw[[dcol, vcol]].copy()
                        _lw.columns = ["日期","_lw"]
                        _lw["日期"] = pd.to_datetime(_lw["日期"], errors="coerce").dt.normalize()
                        val = _lw.loc[_lw["日期"] == day_value, "_lw"].sum()
            except Exception:
                val = float("nan")
            return val

        def _build_rate_dict(ov_slice, lw_value):
            rate_df = (ov_slice
                       .groupby("项目", as_index=False)["产量(kg)"].sum())
            if rate_df.empty or not (pd.notna(lw_value) and lw_value > 0):
                return {}
            rate_df["产成率(小数)"] = rate_df["产量(kg)"] / lw_value
            return {row["项目"]: row["产成率(小数)"] for _, row in rate_df.iterrows()
                    if pd.notna(row["产成率(小数)"]) and row["产成率(小数)"] > 0}

        def _calc_for_day(day_value, need_detail=False):
            ov_day = overview[overview["日期"] == day_value][["项目","产量(kg)","含税金额","含税单价"]].copy()
            if ov_day.empty:
                return {}, {}, [], float("nan"), ov_day, []
            ov_day["项目"] = ov_day["项目"].map(_unify)
            lw_day = _lw_on(day_value)
            if not (pd.notna(lw_day) and float(lw_day) > 0):
                return {}, {}, [], lw_day, ov_day, []
            rate_dict = _build_rate_dict(ov_day, lw_day)
            if not rate_dict:
                return {}, {}, [], lw_day, ov_day, []

            minors_day = pd.DataFrame()
            if minors is not None and not minors.empty:
                try:
                    day_norm = pd.to_datetime(minors["日期"], errors="coerce").dt.normalize()
                    mask = (day_norm == day_value)
                    minors_day = minors.loc[mask].copy()
                except Exception:
                    minors_day = pd.DataFrame()
            if minors_day.empty:
                return {}, {}, [], lw_day, ov_day, []

            code_col = "子类" if "子类" in minors_day.columns else None
            if code_col is None:
                for cand in ["品项","物料号","名称","物料"]:
                    if cand in minors_day.columns:
                        code_col = cand
                        break
            if code_col is None:
                return {}, {}, [], lw_day, ov_day

            minors_day[code_col] = minors_day[code_col].astype(str).str.strip()
            grouped = minors_day.groupby(["部位大类", code_col], as_index=False)["产量(kg)"].sum()
            code_qty = {}
            code_part = {}
            for _, row in grouped.iterrows():
                code_raw = row[code_col]
                restore_part = row["部位大类"]
                qty = float(row["产量(kg)"]) if pd.notna(row["产量(kg)"]) else 0.0
                if qty == 0:
                    continue
                canon_code = _canon_code_val(code_raw)
                targets_raw = restore_mapping.get(canon_code, [])
                if not targets_raw:
                    continue
                targets = [_unify(t) for t in targets_raw if t]
                valid_targets = [t for t in targets if rate_dict.get(t, 0) > 0]
                if not valid_targets:
                    continue
                code_qty[canon_code] = code_qty.get(canon_code, 0.0) + qty
                code_part[canon_code] = restore_part
            bone_rate_total = 0.0
            for code_k, qty_k in code_qty.items():
                if code_part.get(code_k) == "骨架类" and qty_k:
                    bone_rate_total += qty_k / lw_day
            inc_map = {}
            removed_map = {}
            detail_rows = []
            for _, row in grouped.iterrows():
                restore_part = row["部位大类"]
                code_raw = row[code_col]
                qty = float(row["产量(kg)"]) if pd.notna(row["产量(kg)"]) else 0.0
                if qty == 0:
                    continue
                canon_code = _canon_code_val(code_raw)
                targets_raw = restore_mapping.get(canon_code, [])
                if not targets_raw:
                    continue
                targets = [_unify(t) for t in targets_raw if t]
                valid_targets = [t for t in targets if rate_dict.get(t, 0) > 0]
                if not valid_targets:
                    continue
                use_bone_rule = (restore_part == "骨架类")
                sum_target_rates = sum(rate_dict[t] for t in valid_targets)
                if use_bone_rule:
                    adj_rates = {}
                    for t in valid_targets:
                        adj_rate = rate_dict[t]
                        if t == "骨架类":
                            adj_rate = adj_rate - bone_rate_total
                        if adj_rate > 0:
                            adj_rates[t] = adj_rate
                    denom = sum(adj_rates.values())
                    if denom <= 0:
                        continue
                    removed_map[restore_part] = removed_map.get(restore_part, 0.0) + qty
                    for t, adj_rate in adj_rates.items():
                        share = adj_rate / denom
                        inc = qty * share
                        inc_map[t] = inc_map.get(t, 0.0) + inc
                        if need_detail:
                            detail_rows.append({
                                "日期": day_value,
                                "物料号": code_raw,
                                "源产量(kg)": qty,
                                "目标部位": t,
                                "目标产成率(%)": rate_dict[t] * 100.0,
                                "分配比例(%)": share * 100.0,
                                "增量(kg)": inc
                            })
                    continue
                total_rate = sum_target_rates
                if total_rate <= 0:
                    continue
                removed_map[restore_part] = removed_map.get(restore_part, 0.0) + qty
                base_qty = qty / total_rate
                for t in valid_targets:
                    rate_val = rate_dict[t]
                    share = rate_val / total_rate
                    inc = base_qty * rate_val
                    inc_map[t] = inc_map.get(t, 0.0) + inc
                    if need_detail:
                        detail_rows.append({
                            "日期": day_value,
                            "物料号": code_raw,
                            "源产量(kg)": qty,
                            "目标部位": t,
                            "目标产成率(%)": rate_dict[t] * 100.0,
                            "分配比例(%)": share * 100.0,
                            "增量(kg)": inc
                        })
            return inc_map, removed_map, detail_rows, lw_day, ov_day, []

        month_start = pd.Timestamp(ref_day).normalize().replace(day=1)
        days_range = [d for d in _all_days if pd.Timestamp(d) >= month_start and pd.Timestamp(d) <= ref_day]

        # 当日（展示用）
        inc_map_ref, removed_map_ref, detail_ref, lw_val, ov_day_ref, split_detail_ref = _calc_for_day(ref_day, need_detail=True)

        if ov_day_ref.empty:
            st.info("该日没有总览产量数据。")
            st.stop()
        if not (pd.notna(lw_val) and float(lw_val) > 0):
            st.warning("⚠️ 当日没有可用的毛鸡净重，无法计算产成率。")
            st.stop()


        def _ensure_chest(map_obj):
            if "胸类" in map_obj:
                return
            comp = [k for k in ("胸类-胸","胸类-胸皮") if k in map_obj]
            if comp:
                map_obj["胸类"] = float(sum(map_obj[k] for k in comp))

        def _fmt_pct(v):
            if pd.isna(v):
                return ""
            try:
                val = float(v)
            except Exception:
                return str(v)
            # 这里的值已按百分比口径（*100）计算，直接格式化即可
            return f"{val:.2f}%"

        base_summary = ov_day_ref.groupby("项目", as_index=False)["产量(kg)"].sum()
        base_qty_map = {str(row["项目"]): float(row["产量(kg)"]) for _, row in base_summary.iterrows()}
        inc_qty_map = {str(k): float(v) for k, v in inc_map_ref.items()}
        _ensure_chest(base_qty_map)
        _ensure_chest(inc_qty_map)
        show_projects = set(list(base_qty_map.keys()) + list(inc_qty_map.keys()) + list(removed_map_ref.keys()))
        if show_projects:
            row_order = []
            if 'ORDER' in globals():
                row_order = [p for p in ORDER if p in show_projects]
            remainder = [p for p in show_projects if p not in row_order]
            remainder.sort()
            row_order.extend(remainder)

            ov_rows = []
            for proj in row_order:
                orig_qty = base_qty_map.get(proj, 0.0)
                inc_qty = inc_qty_map.get(proj, 0.0) - removed_map_ref.get(proj, 0.0)
                new_qty = orig_qty + inc_qty
                rate_after = (new_qty / lw_val * 100.0) if (pd.notna(lw_val) and lw_val > 0) else np.nan
                ov_rows.append({
                    "项目": proj,
                    "原产量(kg)": float(orig_qty),
                    "调整量(kg)": float(inc_qty),
                    "调整后产量(kg)": float(new_qty),
                    "调整后产成率(%)": rate_after
                })

            summary_df = pd.DataFrame(ov_rows)
            if not summary_df.empty:
                exclude_for_sum = {"胸类","鸡头+鸡脖+骨架"}
                sum_base = summary_df[~summary_df["项目"].isin(exclude_for_sum)]
                total_row = {
                    "项目": "总计",
                    "原产量(kg)": float(sum_base["原产量(kg)"].sum()),
                    "调整量(kg)": float(sum_base["调整量(kg)"].sum()),
                    "调整后产量(kg)": float(sum_base["调整后产量(kg)"].sum()),
                }
                exclude_for_rate = {"胸类","鸡头+鸡脖+骨架"}
                base_for_rate = summary_df[~summary_df["项目"].isin(exclude_for_rate)]["调整后产量(kg)"]
                adj_qty_for_rate = float(base_for_rate.sum()) if not base_for_rate.empty else 0.0
                tot_rate = (adj_qty_for_rate / lw_val * 100.0) if (pd.notna(lw_val) and lw_val > 0) else np.nan
                total_row["调整后产成率(%)"] = tot_rate
                summary_df = pd.concat([summary_df, pd.DataFrame([total_row])], ignore_index=True)
                display_global = summary_df.copy()
                for _c in ["原产量(kg)", "调整量(kg)", "调整后产量(kg)", "调整后产成率(%)"]:
                    if _c in display_global.columns:
                        display_global[_c] = pd.to_numeric(display_global[_c], errors="coerce").round(2)
                display_global["调整后产成率(%)"] = display_global["调整后产成率(%)"].apply(
                    lambda v: "" if pd.isna(v) else f"{float(v):.2f}%"
                )
                export_restore_day = summary_df.copy()
                st.markdown("### 组合还原后产成率总览")
                st.dataframe(display_global, use_container_width=True)

        # —— 当日物料号还原明细（下钻） —— #
        try:
            if detail_ref:
                detail_df = pd.DataFrame(detail_ref)
                if not detail_df.empty:
                    st.markdown("### 物料号还原明细（当日）")
                    st.dataframe(detail_df, use_container_width=True)
        except Exception:
            pass

        # —— 月累计：本月起至所选日，叠加各日调整 —— #
        try:
            month_mask = (overview["日期"] >= month_start) & (overview["日期"] <= ref_day)
            over_month = overview.loc[month_mask].copy()
            # 累计映射的调整量
            inc_month = {}
            removed_month = {}
            for d in days_range:
                inc_d, rem_d, _, _, _, _ = _calc_for_day(d, need_detail=False)
                for k, v in inc_d.items():
                    inc_month[k] = inc_month.get(k, 0.0) + float(v)
                for k, v in rem_d.items():
                    removed_month[k] = removed_month.get(k, 0.0) + float(v)

            if not over_month.empty:
                month_base = over_month.groupby("项目", as_index=False)["产量(kg)"].sum()
                month_qty_map = {str(row["项目"]): float(row["产量(kg)"]) for _, row in month_base.iterrows()}
                _ensure_chest(month_qty_map)
                _ensure_chest(inc_month)

                show_projects_m = set(list(month_qty_map.keys()) + list(inc_month.keys()) + list(removed_month.keys()))
                if show_projects_m:
                    row_order_m = []
                    if 'ORDER' in globals():
                        row_order_m = [p for p in ORDER if p in show_projects_m]
                    remainder_m = [p for p in show_projects_m if p not in row_order_m]
                    remainder_m.sort()
                    row_order_m.extend(remainder_m)

                    lw_month = float("nan")
                    try:
                        if df_lw is not None and not df_lw.empty:
                            dcol = next((c for c in ["日期","交鸡日期","记帐日期","记账日期","凭证日期","过账日期"] if c in df_lw.columns), None)
                            vcol = next((c for c in ["毛鸡净重(kg)","毛鸡净重","净重","净重(kg)"] if c in df_lw.columns), None)
                            if dcol and vcol:
                                _lw = df_lw[[dcol, vcol]].copy()
                                _lw.columns = ["日期","_lw"]
                                _lw["日期"] = pd.to_datetime(_lw["日期"], errors="coerce").dt.normalize()
                                mask_lw = (_lw["日期"] >= month_start) & (_lw["日期"] <= ref_day)
                                lw_month = _lw.loc[mask_lw, "_lw"].sum()
                    except Exception:
                        lw_month = float("nan")

                    month_rows = []
                    for proj in row_order_m:
                        orig_qty = month_qty_map.get(proj, 0.0)
                        adj_qty = inc_month.get(proj, 0.0) - removed_month.get(proj, 0.0)
                        new_qty = orig_qty + adj_qty
                        rate_after = (new_qty / lw_month * 100.0) if (pd.notna(lw_month) and lw_month > 0) else np.nan
                        month_rows.append({
                            "项目": proj,
                            "原产量(kg)": float(orig_qty),
                            "调整量(kg)": float(adj_qty),
                            "调整后产量(kg)": float(new_qty),
                            "调整后产成率(%)": rate_after
                        })

                    month_summary = pd.DataFrame(month_rows)
                    if not month_summary.empty:
                        exclude_for_sum = {"胸类","鸡头+鸡脖+骨架"}
                        sum_base_m = month_summary[~month_summary["项目"].isin(exclude_for_sum)]
                        total_row_m = {
                            "项目": "总计",
                            "原产量(kg)": float(sum_base_m["原产量(kg)"].sum()),
                            "调整量(kg)": float(sum_base_m["调整量(kg)"].sum()),
                            "调整后产量(kg)": float(sum_base_m["调整后产量(kg)"].sum()),
                        }
                        base_for_rate_m = month_summary[~month_summary["项目"].isin(exclude_for_sum)]["调整后产量(kg)"]
                        adj_qty_for_rate_m = float(base_for_rate_m.sum()) if not base_for_rate_m.empty else 0.0
                        tot_rate_m = (adj_qty_for_rate_m / lw_month * 100.0) if (pd.notna(lw_month) and lw_month > 0) else np.nan
                        total_row_m["调整后产成率(%)"] = tot_rate_m
                        month_summary = pd.concat([month_summary, pd.DataFrame([total_row_m])], ignore_index=True)
                        display_month = month_summary.copy()
                        for _c in ["原产量(kg)", "调整量(kg)", "调整后产量(kg)", "调整后产成率(%)"]:
                            if _c in display_month.columns:
                                display_month[_c] = pd.to_numeric(display_month[_c], errors="coerce").round(2)
                        display_month["调整后产成率(%)"] = display_month["调整后产成率(%)"].apply(_fmt_pct)
                        export_restore_month = month_summary.copy()
                        st.markdown("### 组合还原后产成率总览（月累计）")
                        st.dataframe(display_month, use_container_width=True)
        except Exception:
            pass

except Exception as _e:
    st.error(f"自定义部位还原（同日）模块异常：{_e}")

# === 报表下载 ===
try:
    th_cols = ["产量(kg)","销量(kg)","含税金额","原产量(kg)","调整量(kg)","调整后产量(kg)"]
    # 当日总览允许缺失：缺失时仍可下载，导出时自动跳过该分段
    have_all = all([
        export_core_summary is not None and not export_core_summary.empty,
        export_month_cum is not None and not export_month_cum.empty,
        export_restore_day is not None and not export_restore_day.empty,
        export_restore_month is not None and not export_restore_month.empty,
    ])
    if have_all:
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            sheet_name = "报表"
            start_row = 0

            sections = [
                ("核心指标概览", export_core_summary),
                ("当日总览", export_daily_overview),
                ("本月至今累计", export_month_cum),
                ("部位还原后的产成率", export_restore_day),
                ("组合还原后产成率总览（月累计）", export_restore_month),
            ]

            for title, df_sec in sections:
                if df_sec is None or df_sec.empty:
                    continue
                if title == "核心指标概览":
                    start_row += _write_core_summary_excel_section(writer, sheet_name, start_row, df_sec, title)
                    continue
                df_sec.to_excel(writer, sheet_name=sheet_name, index=False, startrow=start_row+1)
                ws = writer.sheets.get(sheet_name)
                if ws is None:
                    ws = writer.book[sheet_name]
                ws.cell(row=start_row+1, column=1, value=title)
                _apply_excel_formats_by_header(
                    ws=ws,
                    header_row=start_row + 2,
                    data_start_row=start_row + 3,
                    data_end_row=start_row + 2 + len(df_sec),
                    col_format_map=_build_excel_col_format_map(df_sec.columns),
                )
                start_row += len(df_sec) + 3  # 标题行 + 数据 + 间隔行

            trend_price = pd.DataFrame()
            trend_rate = pd.DataFrame()
            try:
                sel_dt = pd.to_datetime(sel, errors="coerce").normalize()
            except Exception:
                sel_dt = pd.NaT
            if sel_dt is None or pd.isna(sel_dt):
                try:
                    sel_dt = pd.to_datetime(overview["日期"], errors="coerce").max()
                    if pd.notna(sel_dt):
                        sel_dt = sel_dt.normalize()
                except Exception:
                    sel_dt = pd.NaT
            if pd.notna(sel_dt):
                month_start = sel_dt.replace(day=1)
                ov_days = pd.to_datetime(overview["日期"], errors="coerce").dt.normalize()
                days_in_range = [
                    d.to_pydatetime()
                    for d in pd.date_range(month_start, sel_dt, freq="D")
                ]
                if days_in_range:
                    frames = []
                    for _d in days_in_range:
                        day_df = build_overview_table_for_day(
                            overview,
                            _d,
                            df_lw,
                            total_qty_df=total_qty_df,
                            percent_as_text=False,
                            round_values=False,
                            minors=minors,
                            restore_mapping=restore_mapping_state,
                        )
                        if day_df is None or day_df.empty:
                            continue
                        sub = day_df[["项目", "含税单价", "产成率%"]].copy()
                        sub["日期"] = _d
                        frames.append(sub)
                    if frames:
                        month_trend = pd.concat(frames, ignore_index=True)
                        price_pivot = month_trend.pivot_table(
                            index="项目",
                            columns="日期",
                            values="含税单价",
                            aggfunc="first",
                        )
                        rate_pivot = month_trend.pivot_table(
                            index="项目",
                            columns="日期",
                            values="产成率%",
                            aggfunc="first",
                        )
                        price_pivot = price_pivot.reindex(index=ORDER)
                        rate_pivot = rate_pivot.reindex(index=ORDER)
                        price_pivot = price_pivot.reindex(columns=days_in_range)
                        rate_pivot = rate_pivot.reindex(columns=days_in_range)
                        col_labels = [f"{d.month}月{d.day}日" for d in days_in_range]
                        price_pivot.columns = col_labels
                        rate_pivot.columns = col_labels

                        trend_price = price_pivot.reset_index().rename(columns={"项目":"含税单价"})
                        trend_rate = rate_pivot.reset_index().rename(columns={"项目":"产成率"})

            if (trend_price is not None and not trend_price.empty) or (trend_rate is not None and not trend_rate.empty):
                trend_sheet = "本月趋势"
                if trend_price is not None and not trend_price.empty:
                    trend_price.to_excel(writer, sheet_name=trend_sheet, index=False, startrow=0)
                    ws_t = writer.sheets.get(trend_sheet)
                    _apply_excel_format_by_col_idx(
                        ws=ws_t,
                        col_start=2,
                        col_end=len(trend_price.columns),
                        data_start_row=2,
                        data_end_row=1 + len(trend_price),
                        fmt="0.00",
                    )
                    start_row_trend = len(trend_price) + 2
                else:
                    start_row_trend = 0
                if trend_rate is not None and not trend_rate.empty:
                    trend_rate.to_excel(writer, sheet_name=trend_sheet, index=False, startrow=start_row_trend)
                    ws_t = writer.sheets.get(trend_sheet)
                    _apply_excel_format_by_col_idx(
                        ws=ws_t,
                        col_start=2,
                        col_end=len(trend_rate.columns),
                        data_start_row=start_row_trend + 2,
                        data_end_row=start_row_trend + 1 + len(trend_rate),
                        fmt='0.00"%"',
                    )

            if export_minors is not None and not export_minors.empty:
                export_minors.to_excel(writer, sheet_name="SKU子类明细", index=False)
                ws_m = writer.sheets.get("SKU子类明细")
                _apply_excel_formats_by_header(
                    ws=ws_m,
                    header_row=1,
                    data_start_row=2,
                    data_end_row=1 + len(export_minors),
                    col_format_map=_build_excel_col_format_map(export_minors.columns),
                )
            if export_month_code_detail is not None and not export_month_code_detail.empty:
                export_month_code_detail.to_excel(writer, sheet_name="物料月均价明细", index=False)
                ws_d = writer.sheets.get("物料月均价明细")
                _apply_excel_formats_by_header(
                    ws=ws_d,
                    header_row=1,
                    data_start_row=2,
                    data_end_row=1 + len(export_month_code_detail),
                    col_format_map=_build_excel_col_format_map(export_month_code_detail.columns),
                )

        st.download_button(
            "下载报表（XLSX）",
            data=buffer.getvalue(),
            file_name="报表.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        missing = []
        if export_month_cum is None or export_month_cum.empty:
            missing.append("本月至今累计")
        if export_restore_day is None or export_restore_day.empty:
            missing.append("部位还原后的产成率")
        if export_restore_month is None or export_restore_month.empty:
            missing.append("组合还原后产成率总览（月累计）")
        if export_core_summary is None or export_core_summary.empty:
            missing.append("核心指标概览")
        if missing:
            st.info("请先生成全部数据后再下载报表（缺少：" + "、".join(missing) + "）。")
except Exception as _e:
    st.error(f"报表下载生成失败：{_e}")


